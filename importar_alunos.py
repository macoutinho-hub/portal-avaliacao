"""
Script para importar alunos directamente do Excel para a base de dados.
Execute:  python importar_alunos.py

Aceita o ficheiro 'turmas alunos.xlsx' (ou outro indicado em FICHEIRO).
"""

import sqlite3
import openpyxl
import re
import sys

DATABASE = "portal.db"
FICHEIRO = "turmas alunos.xlsx"   # altere se necessário
ANO_LETIVO = "2025/2026"

def strip_prefix(h):
    return re.sub(r'^[^:]+:\s*', '', h or '').strip().lower()

def col_idx(headers_stripped, names):
    for n in names:
        for i, h in enumerate(headers_stripped):
            if h == n.lower() or h.startswith(n.lower()):
                return i
    return None

def init_db(db):
    db.executescript("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            email TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            nome TEXT NOT NULL,
            turma TEXT,
            role TEXT NOT NULL DEFAULT 'diretor'
        );
        CREATE TABLE IF NOT EXISTS alunos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            numero TEXT NOT NULL,
            nome TEXT NOT NULL,
            turma TEXT NOT NULL,
            ano_letivo TEXT NOT NULL DEFAULT '2025/2026',
            UNIQUE(numero, ano_letivo)
        );
        CREATE TABLE IF NOT EXISTS notas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            aluno_id INTEGER NOT NULL REFERENCES alunos(id) ON DELETE CASCADE,
            disciplina TEXT NOT NULL,
            periodo INTEGER NOT NULL,
            nota REAL,
            observacoes TEXT
        );
    """)
    from werkzeug.security import generate_password_hash
    cur = db.execute("SELECT id FROM users WHERE role='admin'")
    if not cur.fetchone():
        db.execute(
            "INSERT INTO users (email, password, nome, role) VALUES (?,?,?,?)",
            ("admin@escola.pt", generate_password_hash("admin123"), "Administrador", "admin")
        )
    db.commit()

def main():
    ficheiro = sys.argv[1] if len(sys.argv) > 1 else FICHEIRO

    db = sqlite3.connect(DATABASE)
    db.row_factory = sqlite3.Row
    db.execute("PRAGMA foreign_keys = ON")
    init_db(db)

    try:
        wb = openpyxl.load_workbook(ficheiro, data_only=True)
    except FileNotFoundError:
        print(f"Ficheiro não encontrado: {ficheiro}")
        sys.exit(1)

    ws = wb.active
    headers_raw = [str(c.value).strip() if c.value else "" for c in ws[1]]
    headers_stripped = [strip_prefix(h) for h in headers_raw]

    idx_num   = col_idx(headers_stripped, ["numero interno", "numero", "nº", "num"])
    idx_nome  = col_idx(headers_stripped, ["nome", "name"])
    idx_turma = col_idx(headers_stripped, ["turma", "classe"])

    if idx_nome is None or idx_turma is None:
        print("ERRO: Colunas 'Nome' e 'Turma' não encontradas.")
        print("Cabeçalhos detectados:", headers_raw)
        sys.exit(1)

    print(f"Colunas: número={headers_raw[idx_num] if idx_num is not None else 'N/A'} | "
          f"nome={headers_raw[idx_nome]} | turma={headers_raw[idx_turma]}")
    print(f"Ano letivo: {ANO_LETIVO}\n")

    criados = 0
    actualizados = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        nome_val = row[idx_nome] if idx_nome is not None else None
        if not nome_val:
            continue
        turma_val = str(row[idx_turma]).strip() if row[idx_turma] else ""
        num_val   = str(row[idx_num]).strip() if idx_num is not None and row[idx_num] else ""

        cur = db.execute(
            "INSERT OR IGNORE INTO alunos (numero, nome, turma, ano_letivo) VALUES (?,?,?,?)",
            (num_val, str(nome_val).strip(), turma_val, ANO_LETIVO)
        )
        if cur.rowcount == 0:
            db.execute(
                "UPDATE alunos SET nome=?, turma=? WHERE numero=? AND ano_letivo=?",
                (str(nome_val).strip(), turma_val, num_val, ANO_LETIVO)
            )
            actualizados += 1
        else:
            criados += 1

    db.commit()
    db.close()

    print(f"✓ {criados} alunos importados, {actualizados} actualizados.")
    print(f"→ Base de dados: {DATABASE}")

if __name__ == "__main__":
    main()
