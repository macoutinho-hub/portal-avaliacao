# Portal de Avaliação — Colégio Pedro Arrupe
import os
import sqlite3
import secrets
from functools import wraps
from datetime import datetime

from flask import (Flask, render_template, request, redirect, url_for,
                   session, flash, g, jsonify)
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import openpyxl

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", secrets.token_hex(32))

import math as _math
def arred(x):
    """Arredondamento aritmético: .5 arredonda sempre para cima (evita banker's rounding do Python)."""
    return int(_math.floor(x + 0.5))

DATABASE = os.environ.get("DATABASE", "portal.db")
UPLOAD_FOLDER = os.environ.get("UPLOAD_FOLDER", "uploads")
FOTOS_FOLDER  = os.environ.get("FOTOS_FOLDER", os.path.join(os.path.dirname(DATABASE) if os.path.dirname(DATABASE) else ".", "fotos"))
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(FOTOS_FOLDER,  exist_ok=True)

# Inicializar BD ao arrancar (funciona com gunicorn e python app.py)
# Chamado depois de 'app' ser criado, via with app.app_context()
def _auto_init():
    with app.app_context():
        init_db()

import atexit as _atexit

# ─── DB ────────────────────────────────────────────────────────────────────────

def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(DATABASE)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA foreign_keys = ON")
    return g.db

@app.teardown_appcontext
def close_db(e=None):
    db = g.pop("db", None)
    if db:
        db.close()

def init_db():
    db = sqlite3.connect(DATABASE)
    db.row_factory = sqlite3.Row
    db.executescript("""
        CREATE TABLE IF NOT EXISTS users (
            id        INTEGER PRIMARY KEY AUTOINCREMENT,
            email     TEXT UNIQUE NOT NULL,
            password  TEXT NOT NULL,
            nome      TEXT NOT NULL,
            turma     TEXT,          -- NULL = admin
            role      TEXT NOT NULL DEFAULT 'diretor'  -- 'admin' | 'diretor'
        );

        CREATE TABLE IF NOT EXISTS alunos (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            numero     TEXT NOT NULL,
            nome       TEXT NOT NULL,
            turma      TEXT NOT NULL,
            ano_letivo TEXT NOT NULL DEFAULT '2025/2026',
            UNIQUE(numero, ano_letivo)
        );

        CREATE TABLE IF NOT EXISTS notas (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            aluno_id     INTEGER NOT NULL REFERENCES alunos(id) ON DELETE CASCADE,
            disciplina   TEXT NOT NULL,
            periodo      INTEGER NOT NULL,
            nota         REAL,
            observacoes  TEXT
        );

        CREATE TABLE IF NOT EXISTS notas_reuniao (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            aluno_id     INTEGER NOT NULL REFERENCES alunos(id) ON DELETE CASCADE,
            categoria    TEXT NOT NULL,
            texto        TEXT NOT NULL DEFAULT '',
            updated_at   TEXT,
            updated_by   INTEGER REFERENCES users(id),
            UNIQUE(aluno_id, categoria)
        );

        CREATE TABLE IF NOT EXISTS settings (
            key   TEXT PRIMARY KEY,
            value TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS notas_finais (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            aluno_id     INTEGER NOT NULL REFERENCES alunos(id) ON DELETE CASCADE,
            disciplina   TEXT NOT NULL,
            ano_letivo   TEXT NOT NULL,
            cif          REAL,
            exame_f1     REAL,
            exame_f2     REAL,
            cfd          REAL,
            UNIQUE(aluno_id, disciplina, ano_letivo)
        );
    """)
    db.executescript("""
        CREATE TABLE IF NOT EXISTS slides_turma (
            id        INTEGER PRIMARY KEY AUTOINCREMENT,
            turma     TEXT NOT NULL,
            titulo    TEXT NOT NULL DEFAULT '',
            conteudo  TEXT NOT NULL DEFAULT '',
            tipo      TEXT NOT NULL DEFAULT 'texto',  -- 'texto' | 'lista' | 'imagem'
            imagem    TEXT,          -- path relativo ao FOTOS_FOLDER
            ordem     INTEGER NOT NULL DEFAULT 0,
            updated_at TEXT
        );

        CREATE TABLE IF NOT EXISTS auto_avaliacao (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            aluno_id   INTEGER NOT NULL REFERENCES alunos(id) ON DELETE CASCADE,
            disciplina TEXT NOT NULL,
            ano_letivo TEXT NOT NULL,
            valor      INTEGER,
            UNIQUE(aluno_id, disciplina, ano_letivo)
        );

        CREATE TABLE IF NOT EXISTS inscricoes_exame (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            aluno_id     INTEGER NOT NULL REFERENCES alunos(id) ON DELETE CASCADE,
            disciplina   TEXT NOT NULL,
            cod_exame    TEXT NOT NULL,
            interno      TEXT NOT NULL DEFAULT 'N',
            aprovacao    TEXT NOT NULL DEFAULT 'N',
            melhoria     TEXT NOT NULL DEFAULT 'N',
            ano_letivo   TEXT NOT NULL,
            UNIQUE(aluno_id, cod_exame, ano_letivo)
        );
    """)
    db.commit()

    for col_sql in [
        "ALTER TABLE alunos ADD COLUMN bi TEXT",
        "ALTER TABLE notas ADD COLUMN nota_texto TEXT",
        "ALTER TABLE notas ADD COLUMN nivel_curricular INTEGER",
    ]:
        try: db.execute(col_sql); db.commit()
        except Exception: pass

    # Migrar dados existentes: preencher nivel_curricular a partir da turma do aluno
    try:
        db.execute("""
            UPDATE notas SET nivel_curricular = (
                SELECT CAST(SUBSTR(a.turma, 1, 2) AS INTEGER)
                FROM alunos a WHERE a.id = notas.aluno_id
            )
            WHERE nivel_curricular IS NULL
        """)
        db.commit()
    except Exception:
        pass

    # Valores por defeito das settings
    for key, val in [("ano_letivo_atual", "2025/2026"), ("semestre_atual", "2")]:
        db.execute("INSERT OR IGNORE INTO settings (key, value) VALUES (?,?)", (key, val))
    db.commit()
    db.close()
    db.commit()
    # Create default admin if not exists
    cur = db.execute("SELECT id FROM users WHERE role='admin'")
    if not cur.fetchone():
        db.execute(
            "INSERT INTO users (email, password, nome, role) VALUES (?,?,?,?)",
            ("admin@escola.pt", generate_password_hash("admin123"), "Administrador", "admin")
        )
        db.commit()
    db.close()

# ─── Auth helpers ──────────────────────────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        # Verifica o role original (caso esteja em modo de impersonação)
        role_efectivo = session.get("original_role") or session.get("role")
        if role_efectivo != "admin":
            flash("Acesso restrito a administradores.", "danger")
            return redirect(url_for("dashboard"))
        return f(*args, **kwargs)
    return decorated

# ─── Utilitários de turma ─────────────────────────────────────────────────────

import re as _re_global

# Mapeamento código de exame → nome de disciplina na BD
MAPA_COD_EXAME = {
    "639": "Português",
    "635": "Matemática A",
    "735": "Matemática B",
    "835": "Matemática Aplicada Ciências Sociais",
    "702": "Biologia e Geologia",
    "715": "Física e Química A",
    "719": "Geografia A",
    "712": "Economia A",
    "714": "Filosofia",
    "623": "História A",
    "723": "História B",
    "724": "História da Cultura e das Artes",
    "706": "Desenho A",
    "708": "Geometria Descritiva A",
    "734": "Literatura Portuguesa",
    "550": "Líng. Estrang. I - Inglês",
}

# Nomes de colunas usados nas pautas do 11º → nome canónico da BD
# Chaves em minúsculas para comparação case-insensitive
DISC_RENAME_11 = {
    "mat. g (a)":                        "Matemática A",
    "matemática geral (a)":              "Matemática A",
    "mat. g (b)":                        "Matemática B",
    "matemática geral (b)":              "Matemática B",
    "mat. g (macs)":                     "Matemática Aplicada Ciências Sociais",
    "matemática geral (macs)":           "Matemática Aplicada Ciências Sociais",
    "hist. g (a)":                       "História A",
    "história geral (a)":                "História A",
    "hist. g (b)":                       "História B",
    "história geral (b)":                "História B",
    "desenho geral":                     "Desenho A",
    "des. g":                            "Desenho A",
    "des g":                             "Desenho A",
    "desenho g":                         "Desenho A",
}

# Famílias de disciplinas equivalentes (mesma disciplina, nomes diferentes entre anos)
DISC_FAMILIAS = {
    # Variantes da mesma disciplina entre anos (nome canónico = o do ano actual do aluno)
    "Matemática A":                         "família_mat",
    "Matemática Geral":                     "família_mat",
    "Matemática B":                         "família_mat",
    "Matemática Aplicada Ciências Sociais": "família_mat",
    "História A":                           "família_hist",
    "História B":                           "família_hist",
    "História Geral":                       "família_hist",
    "Desenho A":                            "família_des",
    "Desenho Geral":                        "família_des",
    "Inglês":                               "família_ing",
    "Líng. Estrang. I - Inglês":            "família_ing",
}

def nome_match(nome_a, nome_b):
    """
    Verifica se dois nomes correspondem ao mesmo aluno, mesmo com abreviações.
    Ex: 'Diogo Pessoa P. Ferreira' == 'Diogo Pessoa Pereira Ferreira'
    """
    import unicodedata as _uc, re as _re2

    def _norm_palavra(s):
        s = _uc.normalize('NFKD', str(s))
        return s.encode('ascii','ignore').decode().lower().strip('. ')

    def _palavras(nome):
        return [_norm_palavra(p) for p in nome.strip().split() if _norm_palavra(p)]

    pa = _palavras(nome_a)
    pb = _palavras(nome_b)

    if not pa or not pb:
        return False

    # Primeiro e último nome devem coincidir
    if pa[0] != pb[0] or pa[-1] != pb[-1]:
        return False

    # Nomes do meio: comparar par a par, permitindo inicial
    meios_a = pa[1:-1]
    meios_b = pb[1:-1]

    # Se um lado não tem nomes do meio, aceitar
    if not meios_a or not meios_b:
        return True

    # Alinhar pelo menor comprimento
    for ma, mb in zip(meios_a, meios_b):
        if ma == mb:
            continue
        # Verificar se um é inicial do outro
        if len(ma) == 1 and mb.startswith(ma):
            continue
        if len(mb) == 1 and ma.startswith(mb):
            continue
        return False

    return True


def membros_familia(disc):
    """Devolve todos os nomes que pertencem à mesma família que `disc`."""
    fam = DISC_FAMILIAS.get(disc)
    if not fam:
        return [disc]
    return [d for d, f in DISC_FAMILIAS.items() if f == fam]


def encontrar_aluno_id(db, nome_str, turma, ano, numero="", alunos_cache=None, alunos_todos=None):
    """
    Procura aluno_id por: número → nome exacto → nome fuzzy (iniciais).
    alunos_cache: {(nome_norm, turma): id}
    alunos_todos: {nome_norm: [{id, numero, nome_original}]}
    """
    import unicodedata as _uc3, re as _re5

    def _norm(s):
        s = _uc3.normalize('NFKD', str(s or ''))
        s = s.encode('ascii','ignore').decode()
        return _re5.sub(r'\s+', ' ', s).strip().lower()

    nome_n = _norm(nome_str)

    # 1. Por número de processo (mais fiável)
    if numero:
        r = db.execute(
            "SELECT id FROM alunos WHERE numero=? AND ano_letivo=?", (numero, ano)
        ).fetchone()
        if r:
            return r["id"]

    # 2. Por nome normalizado exacto + turma
    if alunos_cache is not None:
        aid = alunos_cache.get((nome_n, turma))
        if aid:
            return aid

    # 3. Por nome normalizado exacto (qualquer turma desse ano)
    r = db.execute(
        "SELECT id FROM alunos WHERE ano_letivo=? AND LOWER(nome)=?", (ano, nome_str.lower())
    ).fetchone()
    if r:
        return r["id"]

    # 4. Fuzzy: nome com iniciais — procurar em todos os alunos do ano
    candidatos = db.execute(
        "SELECT id, nome FROM alunos WHERE ano_letivo=? AND turma=?", (ano, turma)
    ).fetchall()
    for c in candidatos:
        if nome_match(nome_str, c["nome"]):
            return c["id"]

    # 5. Fuzzy sem turma
    if alunos_todos:
        for _, lista in alunos_todos.items():
            for item in lista:
                if nome_match(nome_str, item.get("nome_original", "")):
                    # Verificar se existe registo para o ano
                    r2 = db.execute(
                        "SELECT id FROM alunos WHERE numero=? AND ano_letivo=?",
                        (item.get("numero",""), ano)
                    ).fetchone()
                    if r2:
                        return r2["id"]

    return None


def get_setting(db, key, default=None):
    r = db.execute("SELECT value FROM settings WHERE key=?", (key,)).fetchone()
    return r["value"] if r else default

def ano_letivo_atual(db):
    """Devolve o ano letivo actual das settings (ou o mais recente na BD como fallback)."""
    s = get_setting(db, "ano_letivo_atual")
    if s: return s
    r = db.execute("SELECT ano_letivo FROM alunos ORDER BY ano_letivo DESC LIMIT 1").fetchone()
    return r["ano_letivo"] if r else "2025/2026"

def semestre_atual(db):
    return int(get_setting(db, "semestre_atual") or "2")

def base_turma(turma):
    """Remove sufixo de curso: '12A1 CT' → '12A1', '10A1' → '10A1'."""
    return _re_global.sub(r'\s+\w+$', '', str(turma or "").strip())

def turmas_base_para_sql(turma_base, db, ano):
    """Devolve lista de turmas reais que correspondem à turma base."""
    rows = db.execute(
        "SELECT DISTINCT turma FROM alunos WHERE ano_letivo=? ORDER BY turma", (ano,)
    ).fetchall()
    return [r["turma"] for r in rows if base_turma(r["turma"]) == turma_base]


# ─── Routes ────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    if "user_id" in session:
        return redirect(url_for("dashboard"))
    return redirect(url_for("login"))

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email = request.form["email"].strip().lower()
        password = request.form["password"]
        db = get_db()
        user = db.execute("SELECT * FROM users WHERE email=?", (email,)).fetchone()
        if user and check_password_hash(user["password"], password):
            session.clear()
            session["user_id"] = user["id"]
            session["nome"] = user["nome"]
            session["role"] = user["role"]
            session["turma"] = user["turma"]
            return redirect(url_for("dashboard"))
        flash("Email ou password incorretos.", "danger")
    return render_template("login.html")

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

# ─── Impersonação ──────────────────────────────────────────────────────────────

@app.route("/admin/impersonar/<int:uid>", methods=["POST"])
@login_required
@admin_required
def impersonar(uid):
    db = get_db()
    user = db.execute("SELECT * FROM users WHERE id=?", (uid,)).fetchone()
    if not user:
        flash("Utilizador não encontrado.", "danger")
        return redirect(url_for("admin_panel"))
    if user["role"] == "admin":
        flash("Não é possível personificar outro administrador.", "danger")
        return redirect(url_for("admin_panel"))

    # Guardar sessão original do admin
    session["original_user_id"] = session["user_id"]
    session["original_nome"]    = session["nome"]
    session["original_role"]    = session["role"]
    session["original_turma"]   = session.get("turma", "")

    # Assumir identidade do PT
    session["user_id"] = user["id"]
    session["nome"]    = user["nome"]
    session["role"]    = user["role"]
    session["turma"]   = user["turma"] or ""

    flash(f"A personificar {user['nome']}. Clica em «Sair da personificação» para voltar.", "warning")
    return redirect(url_for("dashboard"))

@app.route("/admin/sair-impersonacao")
@login_required
def sair_impersonacao():
    if "original_user_id" not in session:
        return redirect(url_for("dashboard"))

    # Restaurar sessão original
    session["user_id"] = session.pop("original_user_id")
    session["nome"]    = session.pop("original_nome")
    session["role"]    = session.pop("original_role")
    session["turma"]   = session.pop("original_turma")

    flash("Sessão de personificação terminada.", "success")
    return redirect(url_for("admin_panel"))

@app.route("/alterar-password", methods=["GET", "POST"])
@login_required
def alterar_password():
    if request.method == "POST":
        atual    = request.form.get("password_atual", "")
        nova     = request.form.get("nova_password", "")
        confirma = request.form.get("confirmar_password", "")

        if nova != confirma:
            flash("As passwords não coincidem.", "danger")
            return redirect(url_for("alterar_password"))
        if len(nova) < 6:
            flash("A nova password deve ter pelo menos 6 caracteres.", "danger")
            return redirect(url_for("alterar_password"))

        db = get_db()
        user = db.execute("SELECT * FROM users WHERE id=?", (session["user_id"],)).fetchone()
        if not check_password_hash(user["password"], atual):
            flash("Password actual incorrecta.", "danger")
            return redirect(url_for("alterar_password"))

        db.execute("UPDATE users SET password=? WHERE id=?",
                   (generate_password_hash(nova), session["user_id"]))
        db.commit()
        flash("Password alterada com sucesso.", "success")
        return redirect(url_for("dashboard"))

    return render_template("alterar_password.html")

def calcular_alunos_info(db, turma, ano=None):
    """Devolve lista de alunos com média e nº de negativas para uma turma (base)."""
    if ano is None:
        ano = ano_letivo_atual(db)
    # Suporta turma base ("12A1") ou turma exacta ("12A1 CT")
    turmas_reais = turmas_base_para_sql(turma, db, ano)
    if not turmas_reais:
        turmas_reais = [turma]  # fallback
    placeholders = ",".join("?" * len(turmas_reais))
    alunos = db.execute(
        f"SELECT * FROM alunos WHERE turma IN ({placeholders}) AND ano_letivo=? ORDER BY nome",
        turmas_reais + [ano]
    ).fetchall()
    alunos_info = []
    for a in alunos:
        periodos = db.execute(
            "SELECT DISTINCT periodo FROM notas WHERE aluno_id=? ORDER BY periodo DESC",
            (a["id"],)
        ).fetchall()
        ultimo_periodo = periodos[0]["periodo"] if periodos else None
        media = None
        num_negas = 0
        if ultimo_periodo:
            notas = db.execute(
                "SELECT nota FROM notas WHERE aluno_id=? AND periodo=? AND nota IS NOT NULL",
                (a["id"], ultimo_periodo)
            ).fetchall()
            vals = [n["nota"] for n in notas if n["nota"] is not None]
            media = round(sum(vals) / len(vals), 1) if vals else None
            num_negas = sum(1 for v in vals if v < 10)
        alunos_info.append({
            "id": a["id"], "numero": a["numero"], "nome": a["nome"],
            "turma": a["turma"], "media": media, "periodo": ultimo_periodo,
            "num_negas": num_negas
        })
    return alunos_info

def calcular_stats_turma(db, turma, periodo_sel=None):
    """Calcula médias por disciplina e comparação entre períodos para uma turma (base)."""
    ano = ano_letivo_atual(db)
    turmas_reais = turmas_base_para_sql(turma, db, ano) or [turma]
    ph = ",".join("?" * len(turmas_reais))
    alunos = db.execute(
        f"SELECT id FROM alunos WHERE turma IN ({ph}) AND ano_letivo=?",
        turmas_reais + [ano]
    ).fetchall()
    ids = [a["id"] for a in alunos]
    if not ids:
        return {}, [], None, []

    periodos_disponiveis = [r["periodo"] for r in db.execute(
        f"SELECT DISTINCT periodo FROM notas WHERE aluno_id IN ({','.join('?'*len(ids))}) ORDER BY periodo",
        ids
    ).fetchall()]

    if not periodo_sel or periodo_sel not in periodos_disponiveis:
        periodo_sel = periodos_disponiveis[-1] if periodos_disponiveis else None

    # Médias por disciplina no período seleccionado
    medias_disciplinas = []
    if periodo_sel:
        rows = db.execute(
            f"SELECT disciplina, AVG(nota) as media FROM notas "
            f"WHERE aluno_id IN ({','.join('?'*len(ids))}) AND periodo=? AND nota IS NOT NULL "
            f"GROUP BY disciplina ORDER BY disciplina",
            ids + [periodo_sel]
        ).fetchall()
        medias_disciplinas = [(r["disciplina"], round(r["media"], 1)) for r in rows]

    # Comparação 1º vs 2º semestre
    comparacao = []
    if len(periodos_disponiveis) >= 2:
        p1, p2 = periodos_disponiveis[0], periodos_disponiveis[1]
        discs1 = {r["disciplina"]: r["media"] for r in db.execute(
            f"SELECT disciplina, AVG(nota) as media FROM notas "
            f"WHERE aluno_id IN ({','.join('?'*len(ids))}) AND periodo=? AND nota IS NOT NULL GROUP BY disciplina",
            ids + [p1]
        ).fetchall()}
        discs2 = {r["disciplina"]: r["media"] for r in db.execute(
            f"SELECT disciplina, AVG(nota) as media FROM notas "
            f"WHERE aluno_id IN ({','.join('?'*len(ids))}) AND periodo=? AND nota IS NOT NULL GROUP BY disciplina",
            ids + [p2]
        ).fetchall()}
        for disc in set(discs1) & set(discs2):
            comparacao.append((disc, round(discs1[disc], 1), round(discs2[disc], 1),
                                round(discs2[disc] - discs1[disc], 1)))

    return medias_disciplinas, periodos_disponiveis, periodo_sel, comparacao

@app.route("/dashboard")
@login_required
def dashboard():
    db = get_db()
    if session["role"] == "admin":
        ano = ano_letivo_atual(db)
        # Agrupar por turma base (ex: 12A1 CT + AV + SE → 12A1)
        rows = db.execute(
            "SELECT turma, COUNT(*) as total FROM alunos WHERE ano_letivo=? GROUP BY turma ORDER BY turma",
            (ano,)
        ).fetchall()
        turmas_agrupadas = {}
        for r in rows:
            tb = base_turma(r["turma"])
            turmas_agrupadas[tb] = turmas_agrupadas.get(tb, 0) + r["total"]
        turmas = [{"turma": t, "total": n} for t, n in sorted(turmas_agrupadas.items())]
        return render_template("dashboard_admin.html", turmas=turmas)
    else:
        # turma pode ser "12A1 AV" ou "12A1 AV,12A1 CT,12A1 SE"
        turmas_str = session.get("turma") or ""
        turmas_lista = [t.strip() for t in turmas_str.split(",") if t.strip()]

        # Converter turmas reais para turmas base (ex: "12A1 CT" → "12A1")
        turmas_base_lista = sorted(set(base_turma(t) for t in turmas_lista))
        turma_sel = request.args.get("turma", turmas_base_lista[0] if turmas_base_lista else "")
        if turma_sel not in turmas_base_lista:
            turma_sel = turmas_base_lista[0] if turmas_base_lista else ""
        periodo_sel = request.args.get("periodo", None)
        if periodo_sel:
            try: periodo_sel = int(periodo_sel)
            except: periodo_sel = None

        ano = ano_letivo_atual(db)
        alunos_info = calcular_alunos_info(db, turma_sel, ano) if turma_sel else []
        medias_disc, periodos_disp, periodo_sel, comparacao = calcular_stats_turma(db, turma_sel, periodo_sel)

        kwargs = dict(alunos=alunos_info, turma=turma_sel,
                      medias_disciplinas=medias_disc,
                      periodos_disponiveis=periodos_disp,
                      periodo_sel=periodo_sel,
                      comparacao_periodos=comparacao)
        if len(turmas_base_lista) > 1:
            kwargs["turmas_multiplas"] = turmas_base_lista
        return render_template("dashboard.html", **kwargs)

@app.route("/aluno/<int:aluno_id>")
@login_required
def aluno(aluno_id):
    db = get_db()
    a = db.execute("SELECT * FROM alunos WHERE id=?", (aluno_id,)).fetchone()
    if not a:
        flash("Aluno não encontrado.", "danger")
        return redirect(url_for("dashboard"))
    # Verificar permissão (suporta múltiplas turmas)
    turmas_user = [base_turma(t) for t in (session.get("turma") or "").split(",")]
    if session["role"] != "admin" and base_turma(a["turma"]) not in turmas_user:
        flash("Não tem permissão para ver este aluno.", "danger")
        return redirect(url_for("dashboard"))

    # ── Recolher todos os dados de notas (ano actual + anteriores) ────────────
    todos_alunos_ids = {a["ano_letivo"]: aluno_id}
    if a["numero"]:
        for outro in db.execute(
            "SELECT id, ano_letivo FROM alunos WHERE numero=? AND ano_letivo!=? ORDER BY ano_letivo",
            (a["numero"], a["ano_letivo"])
        ).fetchall():
            todos_alunos_ids[outro["ano_letivo"]] = outro["id"]

    # Organizar por nivel_curricular (10/11/12) — independente do ano letivo
    # Isto permite que uma aluna no 11º tenha disciplinas de 10º e 11º no mesmo ano
    import re as _re_niv
    _m0 = _re_niv.match(r"(\d+)", str(a["turma"] or ""))
    _nivel_atual = int(_m0.group(1)) if _m0 else 11
    _ano_inicio  = int(a["ano_letivo"].split("/")[0])

    def nivel_para_rotulo(nivel):
        return f"{nivel}º Ano"

    # notas_por_nivel: {nivel(10/11/12): {disciplina: {periodo: val}}}
    notas_por_nivel_raw = {}
    nivel_to_aid = {}  # {nivel_int: aluno_id} — para saber qual aluno_id editar/apagar
    for ano, aid in todos_alunos_ids.items():
        # Determinar o nivel base deste registo
        al_info = db.execute("SELECT turma FROM alunos WHERE id=?", (aid,)).fetchone()
        _mn = _re_niv.match(r"(\d+)", str((al_info["turma"] if al_info else "") or ""))
        nivel_base = int(_mn.group(1)) if _mn else _nivel_atual
        nivel_to_aid[nivel_base] = aid  # último vence (ano mais recente por causa do sort)

        rows = db.execute(
            "SELECT disciplina, periodo, nota, nota_texto, nivel_curricular FROM notas WHERE aluno_id=? ORDER BY disciplina, periodo",
            (aid,)
        ).fetchall()
        for r in rows:
            d   = r["disciplina"]
            val = r["nota_texto"] if r["nota_texto"] else r["nota"]
            # nivel: usar campo guardado se existir, senão usar nivel_base do registo
            nivel = r["nivel_curricular"] if r["nivel_curricular"] else nivel_base
            notas_por_nivel_raw.setdefault(nivel, {}).setdefault(d, {})[r["periodo"]] = val
            # garantir que o nivel explícito também tem o aluno_id mapeado
            nivel_to_aid.setdefault(nivel, aid)

    # notas_por_ano (alias) — chave é o rótulo do ano curricular para retrocompatibilidade
    notas_por_ano = {nivel_para_rotulo(n): d for n, d in sorted(notas_por_nivel_raw.items())}
    # Guardar o nível actual para o CIF
    _ano_atual_key = nivel_para_rotulo(_nivel_atual)

    # ── Abreviaturas das disciplinas (apresentação na tabela) ─────────────────
    ABREVIATURAS = {
        "Português":                        "Port",
        "Líng. Estrang. I - Inglês":        "Ing",
        "Inglês":                           "Ing",
        "Filosofia":                        "Filo",
        "Educação Física":                  "Ed. Fís",
        "Religião":                         "Rel",
        "Projeto":                          "Proj",
        "Matemática A":                     "Mat A",
        "Desenho A":                        "Des A",
        "Desenho Geral":                    "Des G",
        "História A":                       "Hist A",
        "História B":                       "Hist B",
        "História Geral":                   "Hist G",
        "Biologia e Geologia":              "Bio Geo",
        "Biologia":                         "Bio",
        "Física e Química A":               "FQ A",
        "Física":                           "Fís",
        "Química":                          "Quím",
        "Geometria Descritiva A":           "GDA",
        "Economia A":                       "Econ A",
        "Economia C":                       "Econ C",
        "Geografia A":                      "Geo A",
        "Matemática B":                     "Mat B",
        "Matemática Geral":                 "Mat G",
        "Matemática Aplicada Ciências Sociais": "MACS",
        "Filosofia A":                      "Filo A",
        "Ciência Política":                 "C. Pol",
        "Psicologia B":                     "Psic B",
        "Aplicações Informáticas B":        "AI B",
        "Oficinas":                         "Ofic",
        "Literatura Portuguesa":            "Lit. P",
        "Alemão":                           "Alem",
        "Espanhol":                         "Esp",
        "Francês":                          "Fr",
        "Hora de PT":                       "PT",
        "Tempo de Trabalho Autónomo":       "TTA",
    }

    # ── Ordem fixa das disciplinas ────────────────────────────────────────────
    ORDEM_TODAS = [
        # Gerais (antes do separador verde)
        "Português", "Líng. Estrang. I - Inglês", "Inglês",
        "Filosofia", "Educação Física", "Religião", "Projeto",
        # Específicas principais (logo após separador)
        "Matemática A", "Desenho A", "Desenho Geral", "História A",
        # Outras específicas
        "Biologia e Geologia", "Biologia",
        "Física e Química A", "Física", "Química",
        "Geometria Descritiva A",
        "Economia A", "Economia C",
        "Geografia A",
        "História Geral",
        "Matemática B", "Matemática Geral", "Matemática Aplicada Ciências Sociais",
        "Filosofia A", "Ciência Política",
        "Psicologia B", "Aplicações Informáticas B", "Oficinas",
        "Literatura Portuguesa", "Alemão", "Espanhol", "Francês",
        # Por último: disciplinas de gestão/apoio
        "Hora de PT", "Tempo de Trabalho Autónomo",
    ]
    N_GERAIS = 7  # índice onde começa o separador (após Projeto)

    todas_set = {d for ano_d in notas_por_ano.values() for d in ano_d}

    # ── Unificar disciplinas equivalentes entre anos ──────────────────────────
    # Usar o nome do ano actual (chave = label de nível, ex: "11º Ano")
    disc_ano_atual_set = set(notas_por_ano.get(_ano_atual_key, {}).keys())

    disc_canonical = {}
    for d in todas_set:
        fam = DISC_FAMILIAS.get(d)
        if not fam:
            disc_canonical[d] = d
            continue
        # 1. Preferir membro da família presente no ano actual
        membros = [m for m in membros_familia(d) if m in disc_ano_atual_set]
        if not membros:
            # 2. Fallback: membro no ano mais recente com dados
            for label_r in sorted(notas_por_ano.keys(), reverse=True):
                membros = [m for m in membros_familia(d) if m in notas_por_ano.get(label_r, {})]
                if membros: break
        disc_canonical[d] = membros[0] if membros else d

    notas_por_ano_unif = {}
    for ano, disc_ano in notas_por_ano.items():
        notas_por_ano_unif[ano] = {}
        for d, periodos in disc_ano.items():
            can = disc_canonical.get(d, d)
            notas_por_ano_unif[ano].setdefault(can, {}).update(periodos)
    notas_por_ano = notas_por_ano_unif
    todas_set = {d for ano_d in notas_por_ano.values() for d in ano_d}

    def _pos_disc(d):
        for i, nome in enumerate(ORDEM_TODAS):
            if d == nome or (len(nome) >= 6 and d.startswith(nome[:6])):
                return i
        return len(ORDEM_TODAS)

    todas_disciplinas = sorted(todas_set, key=_pos_disc)
    gerais_presentes = [d for d in todas_disciplinas if _pos_disc(d) < N_GERAIS]
    separador_idx = len(gerais_presentes) if len(gerais_presentes) < len(todas_disciplinas) else None

    # ── Construir linhas da tabela por nivel_curricular ──────────────────────
    linhas = []

    for rotulo_ano in sorted(notas_por_ano.keys()):  # ex: "10º Ano", "11º Ano"
        disc_ano = notas_por_ano[rotulo_ano]
        periodos = sorted({p for d in disc_ano.values() for p in d})
        e_atual  = (rotulo_ano == _ano_atual_key)
        # aluno_id correcto para este nível (para edição/apagar inline)
        try:
            _nivel_int = int(rotulo_ano.split("º")[0])
        except (ValueError, IndexError):
            _nivel_int = _nivel_atual
        aid_linha = nivel_to_aid.get(_nivel_int, aluno_id)

        for p in periodos:
            notas_linha = {d: disc_ano.get(d, {}).get(p) for d in todas_disciplinas}
            vals_num = [v for v in notas_linha.values() if isinstance(v, (int, float))]
            linhas.append({
                "label":    f"{rotulo_ano} — {p}º Sem.",
                "tipo":     "semestre",
                "atual":    e_atual,
                "notas":    notas_linha,
                "media":    round(sum(vals_num) / len(vals_num), 1) if vals_num else None,
                "aluno_id": aid_linha,
                "periodo":  p,
            })

    # ── Carregar auto-avaliação ───────────────────────────────────────────────
    aa_rows = db.execute(
        "SELECT disciplina, valor FROM auto_avaliacao WHERE aluno_id=? AND ano_letivo=?",
        (aluno_id, a["ano_letivo"])
    ).fetchall()
    aa_map = {r["disciplina"]: r["valor"] for r in aa_rows}
    # Mapear para nomes canónicos
    aa_canonical = {}
    for d, v in aa_map.items():
        can = disc_canonical.get(d, d)
        aa_canonical[can] = v

    # Linha de auto-avaliação (só aparece se existirem valores inseridos)
    if aa_canonical:
        aa_notas = {d: aa_canonical.get(d) for d in todas_disciplinas}
        aa_vals  = [v for v in aa_notas.values() if isinstance(v, (int, float))]
        linhas.append({
            "label": "Auto-Avaliação",
            "tipo":  "auto_av",
            "atual": True,
            "notas": aa_notas,
            "media": round(sum(aa_vals) / len(aa_vals), 1) if aa_vals else None,
        })

    # ── Carregar notas_finais (CIF/Exame/CFD oficiais) ───────────────────────
    # Procurar em todos os registos do mesmo aluno (por número de processo)
    if a["numero"]:
        todos_ids = [r["id"] for r in db.execute(
            "SELECT id FROM alunos WHERE numero=?", (a["numero"],)
        ).fetchall()]
    else:
        todos_ids = [aluno_id]
    placeholders = ",".join("?" * len(todos_ids))
    nf_rows = db.execute(
        f"SELECT disciplina, ano_letivo, cif, exame_f1, exame_f2, cfd FROM notas_finais "
        f"WHERE aluno_id IN ({placeholders})", todos_ids
    ).fetchall()
    # Indexar por disciplina (preferir o ano mais recente)
    nf_cif  = {}
    nf_ex1  = {}
    nf_ex2  = {}
    nf_cfd  = {}
    for r in sorted(nf_rows, key=lambda x: x["ano_letivo"]):
        d = r["disciplina"]
        if r["cif"]  is not None: nf_cif[d]  = r["cif"]
        if r["exame_f1"] is not None: nf_ex1[d] = r["exame_f1"]
        if r["exame_f2"] is not None: nf_ex2[d] = r["exame_f2"]
        if r["cfd"]  is not None: nf_cfd[d]  = r["cfd"]

    # ── CIF: calculado ou oficial se disponível ───────────────────────────────
    cif_notas = {}
    for d in todas_disciplinas:
        if d in nf_cif:
            cif_notas[d] = arred(nf_cif[d])  # oficial → usar directamente
        else:
            # CIF = média do 2º semestre de cada ano (com fallback para 1º sem. se não houver 2º)
            # Para alunos de 11º: 2º sem 10º + 2º sem 11º (ou 1º sem 11º se não houver 2º)
            notas_2s = []
            anos_ordenados = sorted(notas_por_ano.keys())  # ex: ["10º Ano", "11º Ano"]
            for rotulo in anos_ordenados:
                disc_ano = notas_por_ano[rotulo]
                periodos_d = sorted(disc_ano.get(d, {}).keys())
                if not periodos_d:
                    continue
                # Preferir 2º semestre; fallback para o último disponível
                periodo_escolhido = 2 if 2 in periodos_d else periodos_d[-1]
                nota = disc_ano.get(d, {}).get(periodo_escolhido)
                if isinstance(nota, (int, float)):
                    notas_2s.append(nota)
            cif_notas[d] = arred(sum(notas_2s) / len(notas_2s)) if notas_2s else None

    cif_vals = [v for v in cif_notas.values() if v is not None]
    cif_media = arred(sum(cif_vals) / len(cif_vals)) if cif_vals else None

    linhas.append({
        "label": "CIF",
        "tipo": "cif",
        "atual": True,
        "notas": cif_notas,
        "media": cif_media,
    })

    # ── Exame: 2 linhas (1ª/2ª Fase), escala 0-200, só se houver valores ────
    ex1_notas = {d: (round(nf_ex1[d] * 10) if nf_ex1.get(d) is not None else None) for d in todas_disciplinas}
    ex2_notas = {d: (round(nf_ex2[d] * 10) if nf_ex2.get(d) is not None else None) for d in todas_disciplinas}
    ex1_vals = [v for v in ex1_notas.values() if v is not None]
    ex2_vals = [v for v in ex2_notas.values() if v is not None]
    if ex1_vals:
        linhas.append({"label": "Exame 1ª Fase", "tipo": "exame", "escala": 200, "atual": False,
                       "notas": ex1_notas, "media": None})
    if ex2_vals:
        linhas.append({"label": "Exame 2ª Fase", "tipo": "exame", "escala": 200, "atual": False,
                       "notas": ex2_notas, "media": None})

    # exame_notas em escala 0-20 para cálculo do CFD
    exame_notas = {}
    for d in todas_disciplinas:
        f1 = nf_ex1.get(d)
        f2 = nf_ex2.get(d)
        if f1 is not None and f2 is not None:
            exame_notas[d] = max(f1, f2)
        elif f1 is not None:
            exame_notas[d] = f1
        elif f2 is not None:
            exame_notas[d] = f2
        else:
            exame_notas[d] = None

    # ── CFD: oficial ou calculado (7.5×CIF + 2.5×Exame)/10 ──────────────────
    cfd_notas = {}
    for d in todas_disciplinas:
        if d in nf_cfd:
            cfd_notas[d] = arred(nf_cfd[d])  # importado com exames → usar directamente
        elif cif_notas.get(d) is not None and exame_notas.get(d) is not None:
            cfd_notas[d] = arred((7.5 * cif_notas[d] + 2.5 * exame_notas[d]) / 10)
        else:
            cfd_notas[d] = cif_notas.get(d)  # sem exame → CFD = CIF

    cfd_vals = [v for v in cfd_notas.values() if v is not None]
    cfd_media = arred(sum(cfd_vals) / len(cfd_vals)) if cfd_vals else None

    # ── Inscrições de exame (só para 11º e 12º) ──────────────────────────────
    if _nivel_atual in (11, 12):
        insc_rows = db.execute(
            "SELECT disciplina, interno, aprovacao, melhoria FROM inscricoes_exame "
            "WHERE aluno_id=? AND ano_letivo=?",
            (aluno_id, a["ano_letivo"])
        ).fetchall()
        if insc_rows:
            def _tipo_inscricao(interno, aprovacao, melhoria):
                if melhoria == "S": return "M"
                if interno  == "S": return "I"
                if interno  == "N" and (aprovacao == "S" or melhoria != "S"): return "E"
                return "PI"
            insc_map = {}
            for r in insc_rows:
                # mapear para nome canónico se necessário
                d_can = disc_canonical.get(r["disciplina"], r["disciplina"])
                insc_map[d_can] = _tipo_inscricao(r["interno"], r["aprovacao"], r["melhoria"])
            insc_notas = {d: insc_map.get(d) for d in todas_disciplinas}
            linhas.append({
                "label": "Inscrição Exame",
                "tipo": "inscricao",
                "atual": False,
                "notas": insc_notas,
                "media": None,
            })

    linhas.append({
        "label": "CFD",
        "tipo": "cfd",
        "atual": False,
        "notas": cfd_notas,
        "media": cfd_media,
    })

    # ── Resumo para cabeçalho ─────────────────────────────────────────────────
    ultima_linha_sem = next((l for l in reversed(linhas)
                             if l["tipo"] == "semestre" and l["atual"]), None)
    linha_cfd = next((l for l in reversed(linhas) if l["tipo"] == "cfd"), None)
    resumo = None
    if ultima_linha_sem:
        negas = [(d, n) for d, n in ultima_linha_sem["notas"].items() if n is not None and n < 10]
        media_cfd = linha_cfd["media"] if linha_cfd and linha_cfd.get("media") is not None else ultima_linha_sem["media"]
        resumo = {
            "media_atual": media_cfd,
            "num_negas": len(negas),
            "negas": sorted(negas, key=lambda x: x[1]),
        }

    # Verificar se existe foto
    foto_url = None
    if a["numero"]:
        for ext in ("jpg", "jpeg", "png", "webp"):
            if os.path.exists(os.path.join(FOTOS_FOLDER, f"{a['numero']}.{ext}")):
                foto_url = url_for("foto_aluno", numero=a["numero"])
                break

    # Notas de reunião
    notas_reun_rows = db.execute(
        "SELECT categoria, texto, updated_at FROM notas_reuniao WHERE aluno_id=?",
        (aluno_id,)
    ).fetchall()
    notas_reuniao = {r["categoria"]: {"texto": r["texto"], "updated_at": r["updated_at"]}
                     for r in notas_reun_rows}

    # Permissão de edição
    turmas_user = [base_turma(t) for t in (session.get("turma") or "").split(",")]
    pode_editar = session["role"] == "admin" or base_turma(a["turma"]) in turmas_user

    # Lista completa de disciplinas para o modal "Adicionar Semestre"
    todas_disciplinas_possiveis = [d for d in ORDEM_TODAS if d not in ("Hora de PT", "Tempo de Trabalho Autónomo")]

    # ── Navegação: alunos da mesma turma base (ex: 12A2 SE + 12A2 LH) ──────
    turma_base = base_turma(a["turma"])
    turmas_nav = turmas_base_para_sql(turma_base, db, a["ano_letivo"])
    ph = ",".join("?" * len(turmas_nav))
    colegas = db.execute(
        f"SELECT id, nome FROM alunos WHERE turma IN ({ph}) AND ano_letivo=? ORDER BY nome",
        (*turmas_nav, a["ano_letivo"])
    ).fetchall()
    ids_turma = [c["id"] for c in colegas]
    idx_atual = ids_turma.index(aluno_id) if aluno_id in ids_turma else -1
    aluno_anterior = ids_turma[idx_atual - 1] if idx_atual > 0 else None
    aluno_seguinte = ids_turma[idx_atual + 1] if idx_atual >= 0 and idx_atual < len(ids_turma) - 1 else None

    return render_template("aluno.html", aluno=a,
                           todas_disciplinas=todas_disciplinas,
                           todas_disciplinas_possiveis=todas_disciplinas_possiveis,
                           abreviaturas=ABREVIATURAS,
                           separador_idx=separador_idx,
                           linhas=linhas,
                           resumo=resumo,
                           foto_url=foto_url,
                           notas_reuniao=notas_reuniao,
                           categorias_reuniao=CATEGORIAS_REUNIAO,
                           pode_editar=pode_editar,
                           aa_canonical=aa_canonical,
                           colegas=colegas,
                           aluno_anterior=aluno_anterior,
                           aluno_seguinte=aluno_seguinte)

# ─── Auto-avaliação ───────────────────────────────────────────────────────────

@app.route("/aluno/<int:aluno_id>/auto-avaliacao", methods=["POST"])
@login_required
def guardar_auto_avaliacao(aluno_id):
    db = get_db()
    a = db.execute("SELECT * FROM alunos WHERE id=?", (aluno_id,)).fetchone()
    if not a:
        return jsonify({"ok": False, "erro": "Aluno não encontrado"}), 404
    turmas_user = [base_turma(t) for t in (session.get("turma") or "").split(",")]
    if session["role"] != "admin" and base_turma(a["turma"]) not in turmas_user:
        return jsonify({"ok": False, "erro": "Sem permissão"}), 403

    data       = request.get_json()
    disciplina = (data.get("disciplina") or "").strip()
    ano_letivo = (data.get("ano_letivo") or a["ano_letivo"]).strip()
    valor_str  = str(data.get("valor") or "").strip()

    if not disciplina:
        return jsonify({"ok": False, "erro": "Disciplina obrigatória"}), 400

    if valor_str in ("", "-", "—"):
        db.execute("DELETE FROM auto_avaliacao WHERE aluno_id=? AND disciplina=? AND ano_letivo=?",
                   (aluno_id, disciplina, ano_letivo))
        db.commit()
        return jsonify({"ok": True, "valor": None})

    try:
        valor = int(float(valor_str))
        if not (1 <= valor <= 20):
            return jsonify({"ok": False, "erro": "Valor deve ser entre 1 e 20"}), 400
    except ValueError:
        return jsonify({"ok": False, "erro": "Valor inválido"}), 400

    db.execute(
        "INSERT OR REPLACE INTO auto_avaliacao (aluno_id, disciplina, ano_letivo, valor) VALUES (?,?,?,?)",
        (aluno_id, disciplina, ano_letivo, valor)
    )
    db.commit()
    return jsonify({"ok": True, "valor": valor})


@app.route("/aluno/<int:aluno_id>/auto-avaliacao-form", methods=["POST"])
@login_required
def guardar_auto_av_form(aluno_id):
    db = get_db()
    a = db.execute("SELECT * FROM alunos WHERE id=?", (aluno_id,)).fetchone()
    if not a: return redirect(url_for("dashboard"))
    turmas_user = [base_turma(t) for t in (session.get("turma") or "").split(",")]
    if session["role"] != "admin" and base_turma(a["turma"]) not in turmas_user:
        flash("Sem permissão.", "danger")
        return redirect(url_for("aluno", aluno_id=aluno_id))

    n_discs = int(request.form.get("n_discs", 0))
    guardadas = 0
    for i in range(n_discs):
        disc  = request.form.get(f"disc_{i}", "").strip()
        val_s = request.form.get(f"aa_{i}", "").strip()
        if not disc: continue
        if not val_s:
            db.execute("DELETE FROM auto_avaliacao WHERE aluno_id=? AND disciplina=? AND ano_letivo=?",
                       (aluno_id, disc, a["ano_letivo"]))
            continue
        try:
            val = int(float(val_s))
            if 1 <= val <= 20:
                db.execute(
                    "INSERT OR REPLACE INTO auto_avaliacao (aluno_id, disciplina, ano_letivo, valor) VALUES (?,?,?,?)",
                    (aluno_id, disc, a["ano_letivo"], val)
                )
                guardadas += 1
        except ValueError:
            pass

    db.commit()
    if guardadas:
        flash(f"Auto-avaliação guardada ({guardadas} disciplina(s)).", "success")
    return redirect(url_for("aluno", aluno_id=aluno_id))


# ─── Adicionar semestre manual ────────────────────────────────────────────────

@app.route("/aluno/<int:aluno_id>/adicionar-semestre", methods=["POST"])
@login_required
def adicionar_semestre(aluno_id):
    db = get_db()
    a = db.execute("SELECT * FROM alunos WHERE id=?", (aluno_id,)).fetchone()
    if not a:
        flash("Aluno não encontrado.", "danger")
        return redirect(url_for("dashboard"))
    turmas_user = [base_turma(t) for t in (session.get("turma") or "").split(",")]
    if session["role"] != "admin" and base_turma(a["turma"]) not in turmas_user:
        flash("Sem permissão.", "danger")
        return redirect(url_for("dashboard"))

    ano_letivo       = request.form.get("ano_letivo", "").strip()
    semestre         = int(request.form.get("semestre", 1))
    n_discs          = int(request.form.get("n_discs", 0))
    nivel_curricular = request.form.get("nivel_curricular", "")
    try: nivel_curricular = int(nivel_curricular)
    except: nivel_curricular = None

    if not ano_letivo:
        flash("Ano letivo obrigatório.", "danger")
        return redirect(url_for("aluno", aluno_id=aluno_id))

    # Garantir que existe registo do aluno para este ano letivo
    existe = db.execute(
        "SELECT id FROM alunos WHERE numero=? AND ano_letivo=?", (a["numero"], ano_letivo)
    ).fetchone()
    if not existe:
        db.execute(
            "INSERT OR IGNORE INTO alunos (numero, nome, turma, ano_letivo) VALUES (?,?,?,?)",
            (a["numero"], a["nome"], a["turma"], ano_letivo)
        )
        db.commit()
        existe = db.execute(
            "SELECT id FROM alunos WHERE numero=? AND ano_letivo=?", (a["numero"], ano_letivo)
        ).fetchone()

    aluno_id_ano = existe["id"]
    guardadas = 0

    VALORES_TEXTO_S = {"AM", "NE", "NA", "NP", "ND"}

    for i in range(n_discs):
        disc   = request.form.get(f"disc_{i}", "").strip()
        nota_s = request.form.get(f"nota_{i}", "").strip()
        if not disc or not nota_s or nota_s == "—":
            continue

        nota, nota_texto = None, None
        if nota_s.upper() in VALORES_TEXTO_S:
            nota_texto = nota_s.upper()
        else:
            try:
                nota = float(nota_s.replace(",", "."))
                if not (0 <= nota <= 20): continue
            except ValueError:
                continue

        ex = db.execute(
            "SELECT id FROM notas WHERE aluno_id=? AND disciplina=? AND periodo=?",
            (aluno_id_ano, disc, semestre)
        ).fetchone()
        if ex:
            db.execute("UPDATE notas SET nota=?, nota_texto=?, nivel_curricular=? WHERE id=?",
                       (nota, nota_texto, nivel_curricular, ex["id"]))
        else:
            db.execute(
                "INSERT INTO notas (aluno_id, disciplina, periodo, nota, nota_texto, nivel_curricular) VALUES (?,?,?,?,?,?)",
                (aluno_id_ano, disc, semestre, nota, nota_texto, nivel_curricular)
            )
        guardadas += 1

    db.commit()
    flash(f"✓ {guardadas} nota(s) guardada(s) para {ano_letivo} {semestre}º Semestre.", "success")
    return redirect(url_for("aluno", aluno_id=aluno_id))


# ─── Apagar semestre completo ──────────────────────────────────────────────────

@app.route("/aluno/<int:aluno_id>/apagar-semestre", methods=["POST"])
@login_required
def apagar_semestre(aluno_id):
    db = get_db()
    a = db.execute("SELECT * FROM alunos WHERE id=?", (aluno_id,)).fetchone()
    if not a:
        flash("Aluno não encontrado.", "danger")
        return redirect(url_for("dashboard"))
    turmas_user = [base_turma(t) for t in (session.get("turma") or "").split(",")]
    if session["role"] != "admin" and base_turma(a["turma"]) not in turmas_user:
        flash("Sem permissão.", "danger")
        return redirect(url_for("aluno", aluno_id=aluno_id))

    nota_aluno_id = int(request.form.get("nota_aluno_id", aluno_id))
    periodo       = int(request.form.get("periodo", 0))

    # Verificar que nota_aluno_id pertence ao mesmo aluno (mesmo número de processo)
    if a["numero"]:
        valido = db.execute(
            "SELECT id FROM alunos WHERE id=? AND numero=?", (nota_aluno_id, a["numero"])
        ).fetchone()
    else:
        valido = {"id": nota_aluno_id} if nota_aluno_id == aluno_id else None

    if not valido:
        flash("ID de aluno inválido.", "danger")
        return redirect(url_for("aluno", aluno_id=aluno_id))

    result = db.execute(
        "DELETE FROM notas WHERE aluno_id=? AND periodo=?", (nota_aluno_id, periodo)
    )
    db.commit()
    flash(f"✓ {result.rowcount} nota(s) apagada(s).", "success")
    return redirect(url_for("aluno", aluno_id=aluno_id))


@app.route("/aluno/<int:aluno_id>/adicionar-disciplina", methods=["POST"])
@login_required
def adicionar_disciplina(aluno_id):
    db = get_db()
    a = db.execute("SELECT * FROM alunos WHERE id=?", (aluno_id,)).fetchone()
    if not a: return redirect(url_for("dashboard"))
    turmas_user = [base_turma(t) for t in (session.get("turma") or "").split(",")]
    if session["role"] != "admin" and base_turma(a["turma"]) not in turmas_user:
        flash("Sem permissão.", "danger")
        return redirect(url_for("aluno", aluno_id=aluno_id))

    disciplina  = request.form.get("disciplina", "").strip()
    ano_letivo  = request.form.get("ano_letivo", "").strip()
    semestre    = int(request.form.get("semestre", 2))
    nota_str    = request.form.get("nota", "").strip()

    if not disciplina or not ano_letivo:
        flash("Disciplina e ano letivo são obrigatórios.", "danger")
        return redirect(url_for("aluno", aluno_id=aluno_id))

    # Garantir registo do aluno para este ano letivo
    existe = db.execute(
        "SELECT id FROM alunos WHERE numero=? AND ano_letivo=?", (a["numero"], ano_letivo)
    ).fetchone()
    if not existe:
        db.execute(
            "INSERT OR IGNORE INTO alunos (numero, nome, turma, ano_letivo) VALUES (?,?,?,?)",
            (a["numero"], a["nome"], a["turma"], ano_letivo)
        )
        db.commit()
        existe = db.execute(
            "SELECT id FROM alunos WHERE numero=? AND ano_letivo=?", (a["numero"], ano_letivo)
        ).fetchone()

    aid = existe["id"]
    VALORES_TEXTO = {"AM", "NE", "NA", "NP", "ND"}
    nota, nota_texto = None, None
    if nota_str.upper() in VALORES_TEXTO:
        nota_texto = nota_str.upper()
    elif nota_str:
        try:
            nota = float(nota_str.replace(",", "."))
        except ValueError:
            pass

    ex = db.execute(
        "SELECT id FROM notas WHERE aluno_id=? AND disciplina=? AND periodo=?",
        (aid, disciplina, semestre)
    ).fetchone()
    if ex:
        db.execute("UPDATE notas SET nota=?, nota_texto=? WHERE id=?", (nota, nota_texto, ex["id"]))
    else:
        db.execute(
            "INSERT INTO notas (aluno_id, disciplina, periodo, nota, nota_texto) VALUES (?,?,?,?,?)",
            (aid, disciplina, semestre, nota, nota_texto)
        )
    db.commit()
    flash(f"Disciplina '{disciplina}' adicionada.", "success")
    return redirect(url_for("aluno", aluno_id=aluno_id))


# ─── Notas de reunião ─────────────────────────────────────────────────────────

CATEGORIAS_REUNIAO = [
    ("observacoes",   "Observações Gerais",      "bi-journal-text",       "#2563eb"),
    ("preocupacoes",  "Preocupações / Alertas",  "bi-exclamation-triangle","#dc3545"),
    ("positivos",     "Pontos Positivos",         "bi-star",               "#16a34a"),
    ("apoio_caa",     "Medidas de Apoio / CAA",   "bi-life-preserver",     "#9333ea"),
]

@app.route("/aluno/<int:aluno_id>/notas-reuniao", methods=["POST"])
@login_required
def guardar_nota_reuniao(aluno_id):
    db = get_db()
    a = db.execute("SELECT * FROM alunos WHERE id=?", (aluno_id,)).fetchone()
    if not a:
        return jsonify({"ok": False, "erro": "Aluno não encontrado"}), 404

    turmas_user = [base_turma(t) for t in (session.get("turma") or "").split(",")]
    if session["role"] != "admin" and base_turma(a["turma"]) not in turmas_user:
        return jsonify({"ok": False, "erro": "Sem permissão"}), 403

    data = request.get_json()
    categoria = data.get("categoria", "").strip()
    texto     = data.get("texto", "").strip()

    cats_validas = [c[0] for c in CATEGORIAS_REUNIAO] + ['sintese']
    if categoria not in cats_validas:
        return jsonify({"ok": False, "erro": "Categoria inválida"}), 400

    from datetime import datetime as _dt
    agora = _dt.now().strftime("%Y-%m-%d %H:%M")

    existing = db.execute(
        "SELECT id FROM notas_reuniao WHERE aluno_id=? AND categoria=?",
        (aluno_id, categoria)
    ).fetchone()

    if existing:
        db.execute(
            "UPDATE notas_reuniao SET texto=?, updated_at=?, updated_by=? WHERE id=?",
            (texto, agora, session["user_id"], existing["id"])
        )
    else:
        db.execute(
            "INSERT INTO notas_reuniao (aluno_id, categoria, texto, updated_at, updated_by) VALUES (?,?,?,?,?)",
            (aluno_id, categoria, texto, agora, session["user_id"])
        )
    db.commit()
    return jsonify({"ok": True, "updated_at": agora})


# ─── Edição de nota ───────────────────────────────────────────────────────────

@app.route("/aluno/<int:aluno_id>/editar-nota", methods=["POST"])
@login_required
def editar_nota(aluno_id):
    db = get_db()
    a = db.execute("SELECT * FROM alunos WHERE id=?", (aluno_id,)).fetchone()
    if not a:
        return jsonify({"ok": False, "erro": "Aluno não encontrado"}), 404

    turmas_user = [base_turma(t) for t in (session.get("turma") or "").split(",")]
    if session["role"] != "admin" and base_turma(a["turma"]) not in turmas_user:
        return jsonify({"ok": False, "erro": "Sem permissão"}), 403

    data = request.get_json()
    disciplina = (data.get("disciplina") or "").strip()
    periodo    = data.get("periodo")
    nota_str   = str(data.get("nota") or "").strip()

    if not disciplina or periodo is None:
        return jsonify({"ok": False, "erro": "Dados incompletos"}), 400

    # Converter nota
    # Valores especiais de texto (AM, NE, NA, etc.)
    VALORES_TEXTO = {"AM", "NE", "NA", "NP", "ND"}
    nota_texto = None
    nota = None

    if nota_str in ("", "-", "—"):
        pass  # apagar
    elif nota_str.upper() in VALORES_TEXTO:
        nota_texto = nota_str.upper()
    else:
        try:
            nota = float(nota_str.replace(",", "."))
            if not (0 <= nota <= 20):
                return jsonify({"ok": False, "erro": "Nota deve ser entre 0 e 20"}), 400
        except ValueError:
            return jsonify({"ok": False, "erro": "Valor inválido"}), 400

    existing = db.execute(
        "SELECT id FROM notas WHERE aluno_id=? AND disciplina=? AND periodo=?",
        (aluno_id, disciplina, periodo)
    ).fetchone()

    if existing:
        if nota is None and nota_texto is None:
            db.execute("DELETE FROM notas WHERE id=?", (existing["id"],))
        else:
            db.execute("UPDATE notas SET nota=?, nota_texto=? WHERE id=?",
                       (nota, nota_texto, existing["id"]))
    elif nota is not None or nota_texto is not None:
        db.execute(
            "INSERT INTO notas (aluno_id, disciplina, periodo, nota, nota_texto) VALUES (?,?,?,?,?)",
            (aluno_id, disciplina, periodo, nota, nota_texto)
        )

    db.commit()
    display = nota_texto if nota_texto else nota
    return jsonify({"ok": True, "nota": display, "is_texto": nota_texto is not None})


# ─── Fotos ────────────────────────────────────────────────────────────────────

@app.route("/foto/<numero>")
@login_required
def foto_aluno(numero):
    from flask import send_file
    for ext in ("jpg", "jpeg", "png", "webp"):
        path = os.path.join(FOTOS_FOLDER, f"{numero}.{ext}")
        if os.path.exists(path):
            return send_file(path)
    # Foto não encontrada — devolver placeholder SVG
    svg = f'''<svg xmlns="http://www.w3.org/2000/svg" width="80" height="80" viewBox="0 0 80 80">
      <circle cx="40" cy="40" r="40" fill="#2563eb"/>
      <circle cx="40" cy="32" r="14" fill="rgba(255,255,255,.7)"/>
      <ellipse cx="40" cy="72" rx="24" ry="18" fill="rgba(255,255,255,.7)"/>
    </svg>'''
    from flask import Response
    return Response(svg, mimetype="image/svg+xml")


# ─── Admin routes ──────────────────────────────────────────────────────────────

@app.route("/admin")
@login_required
@admin_required
def admin_panel():
    db = get_db()
    users = db.execute("SELECT * FROM users ORDER BY role, turma, nome").fetchall()
    return render_template("admin.html", users=users)

@app.route("/admin/importar-pts", methods=["GET", "POST"])
@login_required
@admin_required
def importar_pts():
    import secrets, string, io, unicodedata as _uc2, re as _re3

    def gerar_password(n=10):
        chars = string.ascii_letters + string.digits
        return "".join(secrets.choice(chars) for _ in range(n))

    credenciais = None

    if request.method == "POST":
        f = request.files.get("ficheiro")
        if not f or not f.filename.endswith((".xlsx", ".xls")):
            flash("Por favor carregue um ficheiro .xlsx.", "danger")
            return redirect(url_for("importar_pts"))

        path = os.path.join(UPLOAD_FOLDER, secure_filename(f.filename))
        f.save(path)
        db = get_db()

        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        headers = [str(c.value or "").strip().lower() for c in ws[1]]

        def ci(names):
            for n in names:
                for i, h in enumerate(headers):
                    if n.lower() in h: return i
            return None

        idx_nome  = ci(["nome"])
        idx_email = ci(["email"])
        idx_turma = ci(["turma"])

        if idx_nome is None or idx_email is None:
            flash("Colunas 'Nome' e 'Email' não encontradas.", "danger")
            return redirect(url_for("importar_pts"))

        criados = []
        ignorados = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            nome  = str(row[idx_nome] or "").strip()
            email = str(row[idx_email] or "").strip().lower()
            turma = str(row[idx_turma] or "").strip() if idx_turma is not None else ""
            if not nome or not email or "@" not in email:
                continue

            # Normalizar turmas: "12A1, 10C2" → "12A1,10C2"
            turma_norm = ",".join(t.strip() for t in turma.split(",") if t.strip())

            existing = db.execute("SELECT id FROM users WHERE email=?", (email,)).fetchone()
            if existing:
                ignorados.append(nome)
                continue

            pwd = gerar_password()
            db.execute(
                "INSERT INTO users (email, password, nome, turma, role) VALUES (?,?,?,?,?)",
                (email, generate_password_hash(pwd), nome, turma_norm, "diretor")
            )
            criados.append({"nome": nome, "email": email, "turma": turma_norm, "password": pwd})

        db.commit()
        flash(f"✓ {len(criados)} PT(s) criado(s). {len(ignorados)} já existiam.", "success")
        credenciais = criados

    return render_template("importar_pts.html", credenciais=credenciais)


@app.route("/admin/criar_utilizador", methods=["POST"])
@login_required
@admin_required
def criar_utilizador():
    email = request.form["email"].strip().lower()
    nome = request.form["nome"].strip()
    turma = request.form["turma"].strip().upper()
    password = request.form["password"]
    role = request.form.get("role", "diretor")
    db = get_db()
    try:
        db.execute(
            "INSERT INTO users (email, password, nome, turma, role) VALUES (?,?,?,?,?)",
            (email, generate_password_hash(password), nome, turma or None, role)
        )
        db.commit()
        flash(f"Utilizador {nome} criado com sucesso.", "success")
    except sqlite3.IntegrityError:
        flash("Email já existe.", "danger")
    return redirect(url_for("admin_panel"))

@app.route("/admin/editar_utilizador/<int:uid>", methods=["POST"])
@login_required
@admin_required
def editar_utilizador(uid):
    db = get_db()
    nome = request.form["nome"].strip()
    turma = request.form["turma"].strip().upper()
    role = request.form.get("role", "diretor")
    nova_pass = request.form.get("nova_password", "").strip()
    if nova_pass:
        db.execute(
            "UPDATE users SET nome=?, turma=?, role=?, password=? WHERE id=?",
            (nome, turma or None, role, generate_password_hash(nova_pass), uid)
        )
    else:
        db.execute(
            "UPDATE users SET nome=?, turma=?, role=? WHERE id=?",
            (nome, turma or None, role, uid)
        )
    db.commit()
    flash("Utilizador atualizado.", "success")
    return redirect(url_for("admin_panel"))

@app.route("/admin/apagar_utilizador/<int:uid>", methods=["POST"])
@login_required
@admin_required
def apagar_utilizador(uid):
    db = get_db()
    db.execute("DELETE FROM users WHERE id=?", (uid,))
    db.commit()
    flash("Utilizador removido.", "success")
    return redirect(url_for("admin_panel"))

@app.route("/admin/importar", methods=["GET", "POST"])
@login_required
@admin_required
def importar_excel():
    if request.method == "POST":
        f = request.files.get("ficheiro")
        ano = request.form.get("ano_letivo", "2025/2026")
        periodo = int(request.form.get("periodo", 1))
        if not f or not f.filename.endswith((".xlsx", ".xls")):
            flash("Por favor carregue um ficheiro Excel (.xlsx).", "danger")
            return redirect(url_for("importar_excel"))

        filename = secure_filename(f.filename)
        path = os.path.join(UPLOAD_FOLDER, filename)
        f.save(path)

        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb.active
            headers = [str(c.value).strip() if c.value else "" for c in ws[1]]

            # Detetar colunas obrigatórias (case-insensitive)
            # Suporta prefixos como "Aluno: Nome", "Aluno: Turma", etc.
            import re as _re
            def _strip_prefix(h):
                return _re.sub(r'^[^:]+:\s*', '', h).strip().lower()

            h_lower   = [h.lower() for h in headers]
            h_stripped = [_strip_prefix(h) for h in headers]

            def col_idx(names):
                for n in names:
                    nl = n.lower()
                    # tenta match exacto primeiro
                    for i, h in enumerate(h_lower):
                        if nl == h:
                            return i
                    # depois sem prefixo
                    for i, h in enumerate(h_stripped):
                        if nl == h or h.startswith(nl):
                            return i
                return None

            idx_num   = col_idx(["numero interno", "numero", "nº", "n.º", "num"])
            idx_nome  = col_idx(["nome", "name"])
            idx_turma = col_idx(["turma", "classe", "class"])

            if idx_nome is None or idx_turma is None:
                flash("Colunas 'Nome' e 'Turma' não encontradas no ficheiro.", "danger")
                return redirect(url_for("importar_excel"))

            # Colunas de disciplinas = todas as restantes (exceto as de controlo)
            skip = {idx_num, idx_nome, idx_turma}
            # Colunas que podem ser observações: terminam em "_obs" ou "observ"
            disc_cols = []
            for i, h in enumerate(headers):
                if i in skip or not h:
                    continue
                disc_cols.append((i, h))

            db = get_db()
            count_alunos = 0
            count_notas = 0

            for row in ws.iter_rows(min_row=2, values_only=True):
                nome_val = row[idx_nome] if idx_nome is not None else None
                if not nome_val:
                    continue
                turma_val = str(row[idx_turma]).strip().upper() if row[idx_turma] else ""
                num_val = str(row[idx_num]).strip() if idx_num is not None and row[idx_num] else ""

                # Inserir/ignorar aluno
                cur = db.execute(
                    "INSERT OR IGNORE INTO alunos (numero, nome, turma, ano_letivo) VALUES (?,?,?,?)",
                    (num_val, str(nome_val).strip(), turma_val, ano)
                )
                if cur.rowcount == 0:
                    # já existe, atualizar nome/turma
                    db.execute(
                        "UPDATE alunos SET nome=?, turma=? WHERE numero=? AND ano_letivo=?",
                        (str(nome_val).strip(), turma_val, num_val, ano)
                    )
                else:
                    count_alunos += 1

                aluno_row = db.execute(
                    "SELECT id FROM alunos WHERE numero=? AND ano_letivo=?", (num_val, ano)
                ).fetchone()
                if not aluno_row:
                    continue
                aluno_id = aluno_row["id"]

                # Determinar nível curricular do aluno para aplicar renomear disciplinas
                _nivel_imp = int(turma_val[:2]) if turma_val[:2].isdigit() else 0

                for col_i, disc_name in disc_cols:
                    val = row[col_i]
                    # Detetar se é observação (coluna seguinte pode ser _obs)
                    is_obs = any(k in disc_name.lower() for k in ["obs", "observ", "nota_text", "descrit"])
                    if is_obs:
                        continue  # tratadas em conjunto com a nota

                    # Normalizar nome da disciplina para alunos de 11º
                    if _nivel_imp == 11:
                        disc_name = DISC_RENAME_11.get(disc_name.strip().lower(), disc_name)

                    nota_num = None
                    try:
                        nota_num = float(val) if val is not None and str(val).strip() != "" else None
                    except (ValueError, TypeError):
                        nota_num = None

                    # Procurar coluna de observações correspondente
                    obs_text = None
                    obs_col_name = disc_name + "_obs"
                    for oi, oh in disc_cols:
                        if oh.lower() == obs_col_name.lower() or (disc_name.lower() in oh.lower() and "obs" in oh.lower()):
                            obs_text = str(row[oi]).strip() if row[oi] else None
                            break

                    # Upsert nota
                    existing = db.execute(
                        "SELECT id FROM notas WHERE aluno_id=? AND disciplina=? AND periodo=?",
                        (aluno_id, disc_name, periodo)
                    ).fetchone()
                    if existing:
                        db.execute(
                            "UPDATE notas SET nota=?, observacoes=? WHERE id=?",
                            (nota_num, obs_text, existing["id"])
                        )
                    else:
                        db.execute(
                            "INSERT INTO notas (aluno_id, disciplina, periodo, nota, observacoes) VALUES (?,?,?,?,?)",
                            (aluno_id, disc_name, periodo, nota_num, obs_text)
                        )
                    count_notas += 1

            db.commit()
            flash(f"Importação concluída: {count_alunos} alunos novos, {count_notas} notas processadas.", "success")

        except Exception as e:
            flash(f"Erro ao processar ficheiro: {e}", "danger")

        return redirect(url_for("importar_excel"))

    return render_template("importar.html")

# ─── Slides de turma ──────────────────────────────────────────────────────────

SLIDES_FOLDER = os.environ.get("SLIDES_FOLDER",
    os.path.join(os.path.dirname(DATABASE) if os.path.dirname(DATABASE) else ".", "slides"))
os.makedirs(SLIDES_FOLDER, exist_ok=True)

def _check_turma_perm(turma):
    tb = base_turma(turma)
    turmas_user = [base_turma(t) for t in (session.get("turma") or "").split(",")]
    return session["role"] == "admin" or tb in turmas_user

@app.route("/turma/<path:turma>/slides", methods=["GET"])
@login_required
def gerir_slides(turma):
    if not _check_turma_perm(turma):
        flash("Sem permissão.", "danger"); return redirect(url_for("dashboard"))
    db = get_db()
    tb = base_turma(turma)
    slides = db.execute(
        "SELECT * FROM slides_turma WHERE turma=? ORDER BY ordem, id", (tb,)
    ).fetchall()
    return render_template("slides.html", turma=tb, slides=slides)

@app.route("/turma/<path:turma>/slides/criar", methods=["POST"])
@login_required
def criar_slide(turma):
    if not _check_turma_perm(turma):
        return redirect(url_for("dashboard"))
    db = get_db()
    tb = base_turma(turma)
    titulo   = request.form.get("titulo", "").strip()
    conteudo = request.form.get("conteudo", "").strip()
    tipo     = request.form.get("tipo", "texto")
    imagem   = None

    f = request.files.get("imagem")
    if f and f.filename:
        ext = os.path.splitext(f.filename)[1].lower()
        nome = f"slide_{tb}_{int(os.times()[4]*1000)}{ext}".replace(" ", "_")
        f.save(os.path.join(SLIDES_FOLDER, nome))
        imagem = nome

    from datetime import datetime as _dt
    max_ord = db.execute("SELECT MAX(ordem) FROM slides_turma WHERE turma=?", (tb,)).fetchone()[0]
    db.execute(
        "INSERT INTO slides_turma (turma, titulo, conteudo, tipo, imagem, ordem, updated_at) VALUES (?,?,?,?,?,?,?)",
        (tb, titulo, conteudo, tipo, imagem, (max_ord or 0) + 1, _dt.now().strftime("%Y-%m-%d %H:%M"))
    )
    db.commit()
    flash("Slide criado.", "success")
    return redirect(url_for("gerir_slides", turma=tb))

@app.route("/turma/<path:turma>/slides/<int:slide_id>/editar", methods=["POST"])
@login_required
def editar_slide(turma, slide_id):
    if not _check_turma_perm(turma):
        return redirect(url_for("dashboard"))
    db = get_db()
    tb = base_turma(turma)
    titulo   = request.form.get("titulo", "").strip()
    conteudo = request.form.get("conteudo", "").strip()
    tipo     = request.form.get("tipo", "texto")
    ordem    = int(request.form.get("ordem", 0))

    f = request.files.get("imagem")
    imagem_update = ""
    if f and f.filename:
        ext = os.path.splitext(f.filename)[1].lower()
        nome = f"slide_{tb}_{slide_id}{ext}".replace(" ", "_")
        f.save(os.path.join(SLIDES_FOLDER, nome))
        imagem_update = nome

    from datetime import datetime as _dt
    if imagem_update:
        db.execute(
            "UPDATE slides_turma SET titulo=?, conteudo=?, tipo=?, imagem=?, ordem=?, updated_at=? WHERE id=? AND turma=?",
            (titulo, conteudo, tipo, imagem_update, ordem, _dt.now().strftime("%Y-%m-%d %H:%M"), slide_id, tb)
        )
    else:
        db.execute(
            "UPDATE slides_turma SET titulo=?, conteudo=?, tipo=?, ordem=?, updated_at=? WHERE id=? AND turma=?",
            (titulo, conteudo, tipo, ordem, _dt.now().strftime("%Y-%m-%d %H:%M"), slide_id, tb)
        )
    db.commit()
    flash("Slide actualizado.", "success")
    return redirect(url_for("gerir_slides", turma=tb))

@app.route("/turma/<path:turma>/slides/<int:slide_id>/apagar", methods=["POST"])
@login_required
def apagar_slide(turma, slide_id):
    if not _check_turma_perm(turma):
        return redirect(url_for("dashboard"))
    db = get_db()
    tb = base_turma(turma)
    db.execute("DELETE FROM slides_turma WHERE id=? AND turma=?", (slide_id, tb))
    db.commit()
    flash("Slide removido.", "success")
    return redirect(url_for("gerir_slides", turma=tb))

@app.route("/slide-img/<filename>")
@login_required
def slide_img(filename):
    from flask import send_file
    path = os.path.join(SLIDES_FOLDER, filename)
    if os.path.exists(path):
        return send_file(path)
    return "", 404


@app.route("/apresentacao/<path:turma>")
@login_required
def apresentacao(turma):
    import json
    db = get_db()

    # Verificar permissão
    turmas_user = [base_turma(t) for t in (session.get("turma") or "").split(",")]
    if session["role"] != "admin" and base_turma(turma) not in turmas_user:
        flash("Sem permissão para esta turma.", "danger")
        return redirect(url_for("dashboard"))

    ano = ano_letivo_atual(db)
    tb = base_turma(turma)
    turmas_reais = turmas_base_para_sql(tb, db, ano) or [turma]
    ph = ",".join("?" * len(turmas_reais))
    alunos = db.execute(
        f"SELECT * FROM alunos WHERE turma IN ({ph}) AND ano_letivo=? ORDER BY nome",
        turmas_reais + [ano]
    ).fetchall()

    alunos_json = []
    for a in alunos:
        # Notas
        rows = db.execute(
            "SELECT disciplina, periodo, nota FROM notas WHERE aluno_id=? ORDER BY disciplina, periodo",
            (a["id"],)
        ).fetchall()

        notas_por_ano = {a["ano_letivo"]: {}}
        for r in rows:
            notas_por_ano[a["ano_letivo"]].setdefault(r["disciplina"], {})[r["periodo"]] = r["nota"]

        # Anos anteriores
        if a["numero"]:
            for outro in db.execute(
                "SELECT id, ano_letivo FROM alunos WHERE numero=? AND ano_letivo!=? ORDER BY ano_letivo",
                (a["numero"], a["ano_letivo"])
            ).fetchall():
                rows_ant = db.execute(
                    "SELECT disciplina, periodo, nota FROM notas WHERE aluno_id=?",
                    (outro["id"],)
                ).fetchall()
                if rows_ant:
                    notas_por_ano[outro["ano_letivo"]] = {}
                    for r in rows_ant:
                        notas_por_ano[outro["ano_letivo"]].setdefault(r["disciplina"], {})[r["periodo"]] = r["nota"]

        # Disciplinas e abreviaturas
        ABREVS = {
            "Português":"Port","Líng. Estrang. I - Inglês":"Ing","Inglês":"Ing",
            "Filosofia":"Filo","Educação Física":"Ed. Fís","Religião":"Rel","Projeto":"Proj",
            "Matemática A":"Mat A","Desenho A":"Des A","Desenho Geral":"Des G","História A":"Hist A",
            "Biologia e Geologia":"Bio Geo","Biologia":"Bio",
            "Física e Química A":"FQ A","Física":"Fís","Química":"Quím",
            "Geometria Descritiva A":"GDA","Economia A":"Econ A","Economia C":"Econ C",
            "Geografia A":"Geo A","História Geral":"Hist B",
            "Matemática B":"Mat B","Matemática Geral":"Mat G","Matemática Aplicada Ciências Sociais":"MACS",
            "Filosofia A":"Filo A","Ciência Política":"C. Pol",
            "Psicologia B":"Psic B","Aplicações Informáticas B":"AI B","Oficinas":"Ofic",
            "Literatura Portuguesa":"Lit. P","Alemão":"Alem","Espanhol":"Esp","Francês":"Fr",
            "Hora de PT":"PT","Tempo de Trabalho Autónomo":"TTA",
        }
        _ORDEM = ["Português","Líng. Estrang. I - Inglês","Inglês","Filosofia",
                  "Educação Física","Religião","Projeto",
                  "Matemática A","Desenho A","Desenho Geral","História A",
                  "Biologia e Geologia","Biologia","Física e Química A","Física","Química",
                  "Geometria Descritiva A","Economia A","Economia C","Geografia A",
                  "História Geral","Matemática B","Matemática Geral","Matemática Aplicada Ciências Sociais",
                  "Filosofia A","Ciência Política","Psicologia B","Aplicações Informáticas B","Oficinas",
                  "Literatura Portuguesa","Alemão","Espanhol","Francês",
                  "Hora de PT","Tempo de Trabalho Autónomo"]
        _N_G = 7
        def _p(d): return next((i for i,n in enumerate(_ORDEM) if d==n or (len(n)>=6 and d.startswith(n[:6]))), len(_ORDEM))

        # ── Canonical renaming por aluno (igual à rota /aluno) ──────────────
        disc_atual_ap = set(notas_por_ano.get(a["ano_letivo"], {}).keys())
        canon_ap = {}
        all_discs_ap = {d for ano_d in notas_por_ano.values() for d in ano_d}
        for d in all_discs_ap:
            fam = DISC_FAMILIAS.get(d)
            if not fam:
                canon_ap[d] = d
                continue
            membros_at = [m for m in membros_familia(d) if m in disc_atual_ap]
            if not membros_at:
                for y in sorted(notas_por_ano.keys(), reverse=True):
                    membros_at = [m for m in membros_familia(d) if m in notas_por_ano.get(y, {})]
                    if membros_at: break
            canon_ap[d] = membros_at[0] if membros_at else d

        notas_por_ano_c = {}
        for y, disc_y in notas_por_ano.items():
            notas_por_ano_c[y] = {}
            for d, periodos in disc_y.items():
                can = canon_ap.get(d, d)
                notas_por_ano_c[y].setdefault(can, {}).update(periodos)
        notas_por_ano = notas_por_ano_c
        # ────────────────────────────────────────────────────────────────────

        todas_set = {d for ano_d in notas_por_ano.values() for d in ano_d}
        todas = sorted(todas_set, key=_p)
        gerais_p = [d for d in todas if _p(d) < _N_G]
        sep_idx = len(gerais_p) if len(gerais_p) < len(todas) else None

        import re as _re
        def ano_turma(t):
            m = _re.match(r"(\d+)", str(t or ""))
            return (m.group(1) + "º Ano") if m else "Ano ?"

        linhas = []
        for ano in sorted(notas_por_ano.keys()):
            disc_ano = notas_por_ano[ano]
            periodos = sorted({p for d in disc_ano.values() for p in d})
            al_ano = db.execute("SELECT turma FROM alunos WHERE numero=? AND ano_letivo=?",
                                (a["numero"], ano)).fetchone()
            t_ano = al_ano["turma"] if al_ano else a["turma"]
            ano_esc = ano_turma(t_ano)

            for p in periodos:
                nl = {d: disc_ano.get(d, {}).get(p) for d in todas}
                vals = [v for v in nl.values() if v is not None]
                linhas.append({"label": f"{ano_esc} — {p}º Sem.", "tipo": "semestre",
                                "atual": ano == a["ano_letivo"], "notas": nl,
                                "media": round(sum(vals)/len(vals),1) if vals else None})

        # CIF: oficial (importado) ou média dos 2ºs semestres — arredondamento aritmético
        cif = {}
        for d in todas:
            if d in ap_cif_of:
                cif[d] = arred(ap_cif_of[d])  # oficial → usar directamente
            else:
                ns = []
                for ano_k, da_k in notas_por_ano.items():
                    pds = sorted(da_k.get(d, {}).keys())
                    if not pds: continue
                    pf = 2 if 2 in pds else pds[-1]
                    n = da_k.get(d, {}).get(pf)
                    if n is not None: ns.append(n)
                cif[d] = arred(sum(ns)/len(ns)) if ns else None
        cv = [v for v in cif.values() if v is not None]
        cm = arred(sum(cv)/len(cv)) if cv else None
        linhas.append({"label":"CIF","tipo":"cif","atual":True,"notas":cif,"media":cm})

        # ── Notas de exame/CIF/CFD da tabela notas_finais ────────────────────
        if a["numero"]:
            ap_ids = [r["id"] for r in db.execute(
                "SELECT id FROM alunos WHERE numero=?", (a["numero"],)
            ).fetchall()]
        else:
            ap_ids = [a["id"]]
        ap_ph = ",".join("?" * len(ap_ids))
        nf_ap = db.execute(
            f"SELECT disciplina, cif, exame_f1, exame_f2, cfd FROM notas_finais WHERE aluno_id IN ({ap_ph})",
            ap_ids
        ).fetchall()
        ap_cif_of, ap_ex1, ap_ex2, ap_cfd_of = {}, {}, {}, {}
        for r in sorted(nf_ap, key=lambda x: x["disciplina"]):
            if r["cif"]      is not None: ap_cif_of[r["disciplina"]] = r["cif"]
            if r["exame_f1"] is not None: ap_ex1[r["disciplina"]]    = r["exame_f1"]
            if r["exame_f2"] is not None: ap_ex2[r["disciplina"]]    = r["exame_f2"]
            if r["cfd"]      is not None: ap_cfd_of[r["disciplina"]] = r["cfd"]

        # Converter 0-20 → 0-200 pontos (inteiro)
        ex1_notas = {d: (round(ap_ex1[d] * 10) if ap_ex1.get(d) is not None else None) for d in todas}
        ex2_notas = {d: (round(ap_ex2[d] * 10) if ap_ex2.get(d) is not None else None) for d in todas}
        ex1_vals = [v for v in ex1_notas.values() if v is not None]
        ex2_vals = [v for v in ex2_notas.values() if v is not None]
        if ex1_vals:
            linhas.append({"label":"Exame 1ª Fase","tipo":"exame","escala":200,"atual":False,
                           "notas":ex1_notas,"media":None})
        if ex2_vals:
            linhas.append({"label":"Exame 2ª Fase","tipo":"exame","escala":200,"atual":False,
                           "notas":ex2_notas,"media":None})

        # ── Inscrições de exame (11º e 12º) ─────────────────────────────────
        import re as _re_ap
        _m_ap = _re_ap.match(r"(\d+)", str(a["turma"] or ""))
        _nivel_ap = int(_m_ap.group(1)) if _m_ap else 0
        if _nivel_ap in (11, 12):
            insc_ap = db.execute(
                "SELECT disciplina, interno, aprovacao, melhoria FROM inscricoes_exame "
                "WHERE aluno_id=? AND ano_letivo=?",
                (a["id"], a["ano_letivo"])
            ).fetchall()
            if insc_ap:
                def _ti(i, ap, m):
                    if m == "S": return "M"
                    if i == "S": return "I"
                    if i == "N": return "E"
                    return "PI"
                insc_map_ap = {}
                for r in insc_ap:
                    d_c = canon_ap.get(r["disciplina"], r["disciplina"])
                    insc_map_ap[d_c] = _ti(r["interno"], r["aprovacao"], r["melhoria"])
                insc_notas_ap = {d: insc_map_ap.get(d) for d in todas}
                # Inserir antes do CFD
                cfd_idx = next((i for i, l in enumerate(linhas) if l["tipo"] == "cfd"), len(linhas))
                linhas.insert(cfd_idx, {
                    "label": "Insc. Exame", "tipo": "inscricao",
                    "atual": False, "notas": insc_notas_ap, "media": None,
                })

        # CFD: oficial (importado com exame) ou calculado (7.5×CIF + 2.5×Exame)/10; sem exame → CFD = CIF
        ap_exame = {}
        for d in todas:
            f1 = ap_ex1.get(d)
            f2 = ap_ex2.get(d)
            if f1 is not None and f2 is not None:
                ap_exame[d] = max(f1, f2)
            elif f1 is not None:
                ap_exame[d] = f1
            elif f2 is not None:
                ap_exame[d] = f2
            else:
                ap_exame[d] = None
        cfd_ap = {}
        for d in todas:
            if d in ap_cfd_of:
                cfd_ap[d] = arred(ap_cfd_of[d])
            elif cif.get(d) is not None and ap_exame.get(d) is not None:
                cfd_ap[d] = arred((7.5 * cif[d] + 2.5 * ap_exame[d]) / 10)
            else:
                cfd_ap[d] = cif.get(d)  # sem exame → CFD = CIF
        cfd_ap_vals = [v for v in cfd_ap.values() if v is not None]
        cfd_ap_media = arred(sum(cfd_ap_vals)/len(cfd_ap_vals)) if cfd_ap_vals else None
        linhas.append({"label":"CFD","tipo":"cfd","atual":False,"notas":cfd_ap,"media":cfd_ap_media})

        # Notas de reunião
        nr_rows = db.execute(
            "SELECT categoria, texto FROM notas_reuniao WHERE aluno_id=?", (a["id"],)
        ).fetchall()
        notas_r = {r["categoria"]: r["texto"] for r in nr_rows}

        # Auto-avaliação (inserir antes do CIF se existir)
        aa_rows_ap = db.execute(
            "SELECT disciplina, valor FROM auto_avaliacao WHERE aluno_id=? AND ano_letivo=?",
            (a["id"], a["ano_letivo"])
        ).fetchall()
        if aa_rows_ap:
            aa_map_ap = {}
            for r in aa_rows_ap:
                can = next((disc_canonical_ap.get(r["disciplina"], r["disciplina"])
                            for disc_canonical_ap in [{}]), r["disciplina"])
                aa_map_ap[r["disciplina"]] = r["valor"]
            aa_notas_ap = {d: aa_map_ap.get(d) for d in todas}
            aa_vals_ap  = [v for v in aa_notas_ap.values() if v is not None]
            # Inserir antes da linha CIF
            cif_idx = next((i for i, l in enumerate(linhas) if l["tipo"] == "cif"), len(linhas))
            linhas.insert(cif_idx, {
                "label": "Auto-Avaliação",
                "tipo":  "auto_av",
                "atual": True,
                "notas": aa_notas_ap,
                "media": round(sum(aa_vals_ap)/len(aa_vals_ap), 1) if aa_vals_ap else None,
            })

        # Negativos última linha semestre
        ul = next((l for l in reversed(linhas) if l["tipo"]=="semestre" and l["atual"]), None)
        negas = []
        if ul:
            negas = [[d, n] for d, n in ul["notas"].items() if n is not None and n < 10]
            negas.sort(key=lambda x: x[1])

        # Foto
        foto_url = None
        if a["numero"]:
            for ext in ("jpg","jpeg","png","webp"):
                if os.path.exists(os.path.join(FOTOS_FOLDER, f"{a['numero']}.{ext}")):
                    foto_url = url_for("foto_aluno", numero=a["numero"])
                    break

        alunos_json.append({
            "id": a["id"], "nome": a["nome"], "numero": a["numero"] or "",
            "turma": a["turma"], "foto_url": foto_url,
            "disciplinas": todas, "abrevs": ABREVS,
            "separador_idx": sep_idx, "linhas": linhas,
            "notas_reuniao": notas_r, "negas": negas,
        })

    slides = db.execute(
        "SELECT id, titulo, conteudo, tipo, imagem FROM slides_turma WHERE turma=? ORDER BY ordem, id",
        (tb,)
    ).fetchall()
    slides_json = json.dumps([{
        "titulo": s["titulo"], "conteudo": s["conteudo"],
        "tipo": s["tipo"],
        "imagem_url": url_for("slide_img", filename=s["imagem"]) if s["imagem"] else None
    } for s in slides])

    return render_template("apresentacao.html",
                           turma=turma,
                           alunos=alunos,
                           alunos_json=json.dumps(alunos_json),
                           slides_json=slides_json)


@app.route("/admin/turma/<path:turma>")
@login_required
@admin_required
def ver_turma(turma):
    db = get_db()
    tb = base_turma(turma)
    ano = ano_letivo_atual(db)
    periodo_sel = request.args.get("periodo", None)
    if periodo_sel:
        try: periodo_sel = int(periodo_sel)
        except: periodo_sel = None
    alunos_info = calcular_alunos_info(db, tb, ano)
    medias_disc, periodos_disp, periodo_sel, comparacao = calcular_stats_turma(db, tb, periodo_sel)
    return render_template("dashboard.html", alunos=alunos_info, turma=tb,
                           medias_disciplinas=medias_disc,
                           periodos_disponiveis=periodos_disp,
                           periodo_sel=periodo_sel,
                           comparacao_periodos=comparacao)

# ─── Importar notas via web (admin) ───────────────────────────────────────────

@app.route("/admin/upload-fotos", methods=["GET", "POST"])
@login_required
@admin_required
def upload_fotos():
    if request.method == "POST":
        f = request.files.get("ficheiro")
        if not f or not f.filename.lower().endswith(".zip"):
            flash("Por favor carregue um ficheiro .zip.", "danger")
            return redirect(url_for("upload_fotos"))

        import zipfile, io

        db = get_db()
        # Obter todos os números de alunos válidos
        numeros = {r["numero"] for r in db.execute("SELECT DISTINCT numero FROM alunos").fetchall() if r["numero"]}

        try:
            zf = zipfile.ZipFile(io.BytesIO(f.read()))
            guardadas = 0
            ignoradas = 0

            for nome in zf.namelist():
                basename = os.path.basename(nome)
                if not basename:
                    continue
                # Extrair número do nome do ficheiro (ex: "3023.jpg" → "3023")
                raiz, ext = os.path.splitext(basename)
                if ext.lower() not in (".jpg", ".jpeg", ".png", ".webp"):
                    continue
                # Normalizar: remover zeros à esquerda para comparação
                raiz_strip = raiz.strip()
                # Verificar se corresponde a um aluno (com ou sem zeros à esquerda)
                match = raiz_strip in numeros or raiz_strip.lstrip("0") in {n.lstrip("0") for n in numeros}
                if not match:
                    ignoradas += 1
                    continue
                # Guardar com o número original do aluno
                numero_aluno = next((n for n in numeros if n.lstrip("0") == raiz_strip.lstrip("0")), raiz_strip)
                dest = os.path.join(FOTOS_FOLDER, f"{numero_aluno}{ext.lower()}")
                with zf.open(nome) as src, open(dest, "wb") as dst:
                    dst.write(src.read())
                guardadas += 1

            flash(f"✓ {guardadas} fotos importadas, {ignoradas} ignoradas (não correspondem a alunos).", "success")
        except Exception as e:
            flash(f"Erro ao processar ZIP: {e}", "danger")

        return redirect(url_for("upload_fotos"))

    # Contar fotos já existentes
    try:
        n_fotos = len([f for f in os.listdir(FOTOS_FOLDER) if f.lower().endswith((".jpg",".jpeg",".png",".webp"))])
    except:
        n_fotos = 0

    return render_template("upload_fotos.html", n_fotos=n_fotos)


# ─── Configurações (admin) ────────────────────────────────────────────────────

@app.route("/admin/configuracoes", methods=["GET", "POST"])
@login_required
@admin_required
def configuracoes():
    db = get_db()
    if request.method == "POST":
        ano = request.form.get("ano_letivo", "").strip()
        sem = request.form.get("semestre", "").strip()
        if ano:
            db.execute("INSERT OR REPLACE INTO settings (key,value) VALUES ('ano_letivo_atual',?)", (ano,))
        if sem in ("1", "2"):
            db.execute("INSERT OR REPLACE INTO settings (key,value) VALUES ('semestre_atual',?)", (sem,))
        db.commit()
        flash("Configurações guardadas.", "success")
        return redirect(url_for("configuracoes"))
    ano = get_setting(db, "ano_letivo_atual", "2025/2026")
    sem = get_setting(db, "semestre_atual", "2")
    return render_template("configuracoes.html", ano_letivo=ano, semestre=sem)


# ─── Importar Pauta (PT / admin) ──────────────────────────────────────────────

def parse_pauta_excel(ws):
    """
    Devolve (turma_str, ano_letivo_str, semestre_int, alunos_list).
    alunos_list: [{numero, nome, {disciplina: {cf, cif, ce, cfd}}}]
    """
    import re as _r
    rows = list(ws.iter_rows(values_only=True))

    # Detectar linha de disciplinas (tem 'PORT' ou similar)
    disc_row_idx = None
    for i, row in enumerate(rows):
        vals = [str(v or "") for v in row]
        if any("PORT" in v or "FILO" in v or "FÍSICA" in v for v in vals):
            disc_row_idx = i
            break
    if disc_row_idx is None:
        return None, None, None, []

    disc_row = rows[disc_row_idx]
    sub_row  = rows[disc_row_idx + 1] if disc_row_idx + 1 < len(rows) else []

    # Mapa: disc_name → {CF: col, CIF: col, CE: col, CFD: col}
    disc_map = {}
    skip_cols = {1, 3}  # Matrícula, Nome
    for ci, v in enumerate(disc_row):
        if not v or not str(v).strip(): continue
        s = str(v).strip()
        if any(k in s for k in ["N.º", "Nome", "Averb", "Afixado"]):
            continue
        # É uma disciplina
        # Encontrar CF, CIF, CE, CFD nas colunas seguintes
        disc_map[s] = {}
        for offset in range(0, 20):
            if ci + offset >= len(sub_row): break
            sv = str(sub_row[ci + offset] or "").strip()
            if sv in ("CF", "CIF", "CE", "CFD") and sv not in disc_map[s]:
                disc_map[s][sv] = ci + offset

    # Turma e ano letivo a partir dos metadados
    turma_str = ""
    ano_letivo_str = ""
    semestre_int = 1
    for row in rows[:disc_row_idx]:
        for v in row:
            if not v: continue
            s = str(v)
            m = _r.search(r"Turma:\s*(\w+)", s)
            if m: turma_str = m.group(1).strip()
            m = _r.search(r"(\d{4})\s*/\s*(\d{4})", s)
            if m: ano_letivo_str = f"{m.group(1)}/{m.group(2)}"
            m = _r.search(r"(\d)[.ºo]\s*semestre", s, _r.I)
            if m: semestre_int = int(m.group(1))

    # Encontrar colunas de matrícula e nome
    col_num  = next((ci for ci, v in enumerate(disc_row) if v and "N.º" in str(v)), 1)
    col_nome = next((ci for ci, v in enumerate(disc_row) if v and "Nome" in str(v)), 3)

    # Ler dados dos alunos
    alunos = []
    for row in rows[disc_row_idx + 2:]:
        num  = str(row[col_num] or "").strip() if col_num < len(row) else ""
        nome = str(row[col_nome] or "").strip() if col_nome < len(row) else ""
        if not num or not nome or not num.isdigit():
            continue
        notas = {}
        for disc, cols in disc_map.items():
            d = {}
            for field, ci in cols.items():
                raw = row[ci] if ci < len(row) else None
                if raw is None: continue
                s = str(raw).strip().replace(" ", "")
                if s in ("-", "", "AM", "NA", "NE", "—"): continue
                try: d[field] = float(s)
                except: pass
            if d: notas[disc] = d
        alunos.append({"numero": num, "nome": nome, "notas": notas})

    return turma_str, ano_letivo_str, semestre_int, alunos


@app.route("/turma/<path:turma>/importar-pauta", methods=["GET", "POST"])
@login_required
def importar_pauta(turma):
    db = get_db()
    tb = base_turma(turma)
    turmas_user = [base_turma(t) for t in (session.get("turma") or "").split(",")]
    if session["role"] != "admin" and tb not in turmas_user:
        flash("Sem permissão.", "danger")
        return redirect(url_for("dashboard"))

    ano_conf = ano_letivo_atual(db)
    sem_conf = semestre_atual(db)

    if request.method == "POST":
        f = request.files.get("ficheiro")
        if not f or not f.filename.endswith((".xlsx", ".xls")):
            flash("Carregue um ficheiro .xlsx.", "danger")
            return redirect(url_for("importar_pauta", turma=turma))

        path = os.path.join(UPLOAD_FOLDER, secure_filename(f.filename))
        f.save(path)

        wb  = openpyxl.load_workbook(path, data_only=True)
        ws  = wb.active
        turma_doc, ano_doc, sem_doc, alunos_doc = parse_pauta_excel(ws)

        # Usar semestre/ano da pauta se detectados, senão usar configuração
        ano_usar = ano_doc or ano_conf
        sem_usar = sem_doc or sem_conf

        # Cache alunos da turma
        turmas_reais = turmas_base_para_sql(tb, db, ano_usar) or [turma]
        ph = ",".join("?" * len(turmas_reais))
        alunos_db = db.execute(
            f"SELECT id, numero, nome FROM alunos WHERE turma IN ({ph}) AND ano_letivo=?",
            turmas_reais + [ano_usar]
        ).fetchall()

        import unicodedata as _uc3, re as _re4
        def _norm(s):
            s = _uc3.normalize("NFKD", str(s or ""))
            s = s.encode("ascii","ignore").decode()
            return _re4.sub(r"\s+"," ",s).strip().lower()

        cache_num  = {a["numero"]: a["id"] for a in alunos_db}
        cache_nome = {_norm(a["nome"]): a["id"] for a in alunos_db}

        total_notas = 0
        nao_enc = []

        for al in alunos_doc:
            # 1. Por número; 2. nome exacto; 3. nome fuzzy
            aluno_id = cache_num.get(al["numero"]) or cache_nome.get(_norm(al["nome"]))
            if aluno_id is None:
                for a_db in alunos_db:
                    if nome_match(al["nome"], a_db["nome"]):
                        aluno_id = a_db["id"]; break
            if aluno_id is None:
                nao_enc.append(al["nome"])
                continue

            # Nível curricular da turma (ex: "11D1" → 11)
            _nivel_pauta = int(tb[:2]) if tb[:2].isdigit() else 0

            for disc_abrev, campos in al["notas"].items():
                disc_nome = MAPA_DISC_GLOBAL.get(disc_abrev, disc_abrev.rstrip(" (a)(b)(macs)").strip())
                # Para turmas de 11º, DES.G é sempre Desenho A
                if _nivel_pauta == 11 and disc_nome == "Desenho Geral":
                    disc_nome = "Desenho A"

                cf  = campos.get("CF")
                cif = campos.get("CIF")
                ce  = campos.get("CE")
                cfd = campos.get("CFD")

                # Importar CF como nota do semestre
                if cf is not None:
                    # Apagar registos de disciplinas equivalentes (mesma família) para o mesmo período
                    for _membro in membros_familia(disc_nome):
                        if _membro != disc_nome:
                            db.execute(
                                "DELETE FROM notas WHERE aluno_id=? AND disciplina=? AND periodo=?",
                                (aluno_id, _membro, sem_usar)
                            )
                    ex = db.execute(
                        "SELECT id FROM notas WHERE aluno_id=? AND disciplina=? AND periodo=?",
                        (aluno_id, disc_nome, sem_usar)
                    ).fetchone()
                    if ex:
                        db.execute("UPDATE notas SET nota=? WHERE id=?", (cf, ex["id"]))
                    else:
                        db.execute(
                            "INSERT INTO notas (aluno_id, disciplina, periodo, nota) VALUES (?,?,?,?)",
                            (aluno_id, disc_nome, sem_usar, cf)
                        )
                    total_notas += 1

                # Importar CIF/CE/CFD em notas_finais se presentes
                if any(v is not None for v in [cif, ce, cfd]):
                    ex = db.execute(
                        "SELECT id FROM notas_finais WHERE aluno_id=? AND disciplina=? AND ano_letivo=?",
                        (aluno_id, disc_nome, ano_usar)
                    ).fetchone()
                    ce_conv = round(ce / 10, 1) if ce and ce > 20 else ce
                    if ex:
                        db.execute(
                            "UPDATE notas_finais SET cif=?, exame_f1=?, cfd=? WHERE id=?",
                            (round(cif) if cif else None, ce_conv, round(cfd) if cfd else None, ex["id"])
                        )
                    else:
                        db.execute(
                            "INSERT INTO notas_finais (aluno_id, disciplina, ano_letivo, cif, exame_f1, cfd) VALUES (?,?,?,?,?,?)",
                            (aluno_id, disc_nome, ano_usar, round(cif) if cif else None, ce_conv, round(cfd) if cfd else None)
                        )

        db.commit()
        msg = f"✓ {total_notas} notas importadas da pauta ({turma_doc or tb}, {ano_usar} {sem_usar}º S.)."
        if nao_enc:
            msg += f" {len(nao_enc)} aluno(s) não encontrado(s)."
        flash(msg, "success" if not nao_enc else "warning")
        return redirect(url_for("ver_turma", turma=tb))

    return render_template("importar_pauta.html", turma=tb,
                           ano_conf=ano_conf, sem_conf=sem_conf)


# Mapeamento global de abreviaturas de disciplinas (pauta → nome completo)
MAPA_DISC_GLOBAL = {
    "PORT.": "Português", "FILO.": "Filosofia",
    "ED.FÍSICA": "Educação Física", "RELIGIÃO": "Religião",
    "LE I-ING.": "Líng. Estrang. I - Inglês",
    "MAT. G (a)": "Matemática A",
    "MAT. G (b)": "Matemática B",
    "MAT. G (macs)": "Matemática Aplicada Ciências Sociais",
    "BIO.GEO.": "Biologia e Geologia", "FÍS.QUÍM.A": "Física e Química A",
    "GEOG.A": "Geografia A",
    "HIST. G (a)": "História A",
    "HIST. G (b)": "História B",
    "HIST.A": "História A",
    "DES.G": "Desenho Geral",  # substituído por Desenho A em turmas de 11º (ver importar_pauta)
    "DES.A": "Desenho A",
    "ECON.A": "Economia A", "GEO.DESC.A": "Geometria Descritiva A",
    "PROJ.": "Projeto", "PT": "Hora de PT",
}


# ─── Importar BI dos alunos (Dados Alunos) ───────────────────────────────────

@app.route("/admin/importar-bi", methods=["GET", "POST"])
@login_required
@admin_required
def importar_bi():
    db = get_db()
    if request.method == "POST":
        f = request.files.get("ficheiro")
        if not f or not f.filename.endswith((".xlsx", ".xls")):
            flash("Por favor carregue um ficheiro .xlsx.", "danger")
            return redirect(url_for("importar_bi"))

        path = os.path.join(UPLOAD_FOLDER, secure_filename(f.filename))
        f.save(path)

        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active

        # Detectar linha de cabeçalho (tem "Nº BI" e "N.º Processo")
        header_idx = None
        headers = []
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True)):
            vals = [str(v or "").strip() for v in row]
            if any("BI" in v for v in vals) and any("Processo" in v for v in vals):
                header_idx = i + 1
                headers = vals
                break
        if header_idx is None:
            flash("Cabeçalho não encontrado (esperado: 'Nº BI' e 'N.º Processo').", "danger")
            return redirect(url_for("importar_bi"))

        def ci(names):
            for n in names:
                for i, h in enumerate(headers):
                    if n.lower() in h.lower(): return i
            return None

        idx_bi      = ci(["nº bi", "bi"])
        idx_numero  = ci(["n.º processo", "processo", "numero interno", "n.º aluno"])

        if idx_bi is None or idx_numero is None:
            flash("Colunas 'Nº BI' e 'N.º Processo' não encontradas.", "danger")
            return redirect(url_for("importar_bi"))

        atualizados = 0
        nao_enc = 0
        for row in ws.iter_rows(min_row=header_idx + 1, values_only=True):
            bi_raw  = str(row[idx_bi] or "").strip() if idx_bi < len(row) else ""
            num_raw = str(row[idx_numero] or "").strip() if idx_numero < len(row) else ""
            if not bi_raw or not num_raw or bi_raw == "None": continue
            # Normalizar BI: só dígitos, primeiros 8
            import re as _re_bi
            bi = _re_bi.sub(r"[^0-9]", "", bi_raw)[:8]
            if not bi: continue
            r = db.execute(
                "UPDATE alunos SET bi=? WHERE numero=? AND (bi IS NULL OR bi='')",
                (bi, num_raw)
            )
            if r.rowcount > 0:
                atualizados += 1
            else:
                # Verificar se o aluno existe
                existe = db.execute("SELECT id FROM alunos WHERE numero=?", (num_raw,)).fetchone()
                if not existe: nao_enc += 1

        db.commit()
        msg = f"✓ {atualizados} BI(s) actualizados."
        if nao_enc: msg += f" {nao_enc} número(s) não encontrados."
        flash(msg, "success" if not nao_enc else "warning")
        return redirect(url_for("importar_bi"))

    n = db.execute("SELECT COUNT(*) FROM alunos WHERE bi IS NOT NULL AND bi != ''").fetchone()[0]
    return render_template("importar_bi.html", n_com_bi=n)


# ─── Ferramenta de auditoria de notas em falta (admin) ────────────────────────

@app.route("/admin/auditoria-notas")
@login_required
@admin_required
def auditoria_notas():
    db = get_db()
    ano = ano_letivo_atual(db)

    # Alunos do 11º e 12º ano actual
    rows = db.execute(
        "SELECT id, numero, nome, turma FROM alunos WHERE ano_letivo=? ORDER BY turma, nome",
        (ano,)
    ).fetchall()

    resultados = []
    for a in rows:
        tb = base_turma(a["turma"])
        # Detectar ano escolar
        import re as _r5
        m = _r5.match(r"(\d+)", tb)
        ano_esc = int(m.group(1)) if m else 0
        if ano_esc not in (11, 12):
            continue

        # Semestres esperados de anos anteriores
        anos_esperados = []
        if ano_esc == 11:
            anos_esperados = [(ano.split("/")[0][:3] + str(int(ano.split("/")[0][-1]) - 1) + "/" + ano.split("/")[0], 1),
                             (ano.split("/")[0][:3] + str(int(ano.split("/")[0][-1]) - 1) + "/" + ano.split("/")[0], 2)]
            # 10º ano = ano letivo anterior
            ano_ant = f"{int(ano.split('/')[0])-1}/{int(ano.split('/')[1])-1}"
            anos_esperados = [(ano_ant, 1), (ano_ant, 2)]
        elif ano_esc == 12:
            ano_ant1 = f"{int(ano.split('/')[0])-2}/{int(ano.split('/')[1])-2}"
            ano_ant2 = f"{int(ano.split('/')[0])-1}/{int(ano.split('/')[1])-1}"
            anos_esperados = [(ano_ant1, 1), (ano_ant1, 2), (ano_ant2, 1), (ano_ant2, 2)]

        # Verificar quais anos/semestres têm notas
        falta = []
        for ano_esp, sem_esp in anos_esperados:
            count = db.execute(
                """SELECT COUNT(*) FROM notas n
                   JOIN alunos al ON al.id = n.aluno_id
                   WHERE al.numero=? AND al.ano_letivo=? AND n.periodo=?""",
                (a["numero"], ano_esp, sem_esp)
            ).fetchone()[0]
            if count == 0:
                falta.append(f"{ano_esp} {sem_esp}ºS")

        if falta:
            resultados.append({
                "id": a["id"], "nome": a["nome"], "turma": a["turma"],
                "ano_esc": f"{ano_esc}º Ano", "falta": falta
            })

    return render_template("auditoria_notas.html", resultados=resultados, ano=ano)


@app.route("/admin/importar-aludisc", methods=["GET", "POST"])
@login_required
@admin_required
def importar_aludisc():
    """Importa ficheiro AluDisc (CIF, Exame, CFD por BI e código de disciplina)."""

    # Mapeamento de códigos internos → nomes de disciplinas
    MAPA_CODIGOS = {
        "N014": "Biologia e Geologia", "N030": "Desenho A",
        "N040": "Economia A",          "N044": "Economia C",
        "N046": "Religião",            "N048": "Educação Física",
        "N096": "Filosofia",           "N100": "Física e Química A",
        "N104": "Biologia",            "N110": "Geografia A",
        "N118": "Geometria Descritiva A", "N128": "História da Cultura e das Artes",
        "N130": "História A",          "N132": "História Geral",
        "N156": "Matemática Aplicada Ciências Sociais",
        "N162": "Matemática A",        "N164": "Matemática Geral",
        "N174": "Literatura Portuguesa","N186": "Português",
        "N220": "Filosofia A",         "N222": "Ciência Política",
        "N230": "Psicologia B",        "N304": "Líng. Estrang. I - Inglês",
        "N314": "Alemão",              "N332": "Espanhol",
        "N334": "Francês",
        "N009": "Aplicações Informáticas B", "N018": "Desenho Geral",
        "N022": "Física",              "N098": "Química",
    }

    if request.method == "POST":
        f = request.files.get("ficheiro")
        if not f or not f.filename.endswith((".xlsx", ".xls", ".txt")):
            flash("Por favor carregue um ficheiro .xlsx ou .txt.", "danger")
            return redirect(url_for("importar_aludisc"))

        path = os.path.join(UPLOAD_FOLDER, secure_filename(f.filename))
        f.save(path)

        # ── Formato TXT de largura fixa (AluDisc) ──────────────────────────
        if f.filename.endswith(".txt"):
            db = get_db()
            bi_map = {}
            for a in db.execute("SELECT id, bi, ano_letivo FROM alunos WHERE bi IS NOT NULL AND bi != ''").fetchall():
                bi_map.setdefault(a["bi"], {})[a["ano_letivo"]] = a["id"]

            import re as _re_txt
            importados = 0
            sem_bi_txt = set()

            def _parse_int(s):
                s = s.strip()
                if not s: return None
                try: return int(s)
                except: return None

            with open(path, encoding="latin-1") as fh:
                for line in fh:
                    if len(line) < 39: continue
                    # Formato AluDisc (posições 1-based):
                    # 1-8=BI, 9=dígito controlo, 10-13=cod disciplina,
                    # 14-17=ano, 18-19=CF10, 20-21=CF11, 22-23=CF12,
                    # 24-25=CIF, 26-27=estado, 28-30=Ex1(0-200),
                    # 31-33=Ex2(0-200), 34-35=CFD, 36-37=CFDa, 38=extracurr
                    bi      = line[1:9].strip()
                    cod     = line[10:14].strip()
                    ano_n   = _parse_int(line[14:18])
                    cif_raw = _parse_int(line[24:26])
                    ex1_raw = _parse_int(line[28:31])
                    ex2_raw = _parse_int(line[31:34])
                    cfd_raw = _parse_int(line[34:36]) or _parse_int(line[36:38])  # CFD; fallback CFDa

                    if not bi or not cod or not ano_n: continue

                    ano_letivo = f"{ano_n-1}/{ano_n}"
                    disc_nome  = MAPA_CODIGOS.get(cod, cod)

                    # Converter exames de 0-200 para 0-20
                    ex1 = round(ex1_raw / 10, 1) if ex1_raw is not None and ex1_raw > 0 else None
                    ex2 = round(ex2_raw / 10, 1) if ex2_raw is not None and ex2_raw > 0 else None

                    # Encontrar aluno pelo BI
                    aluno_id = None
                    if bi in bi_map:
                        anos = bi_map[bi]
                        aluno_id = anos.get(ano_letivo) or anos.get(max(anos.keys()))
                    if aluno_id is None:
                        sem_bi_txt.add(bi); continue

                    ex_db = db.execute(
                        "SELECT id FROM notas_finais WHERE aluno_id=? AND disciplina=? AND ano_letivo=?",
                        (aluno_id, disc_nome, ano_letivo)
                    ).fetchone()
                    if ex_db:
                        db.execute(
                            "UPDATE notas_finais SET cif=?, exame_f1=?, exame_f2=?, cfd=? WHERE id=?",
                            (cif_raw, ex1, ex2, cfd_raw, ex_db["id"])
                        )
                    else:
                        db.execute(
                            "INSERT INTO notas_finais (aluno_id, disciplina, ano_letivo, cif, exame_f1, exame_f2, cfd) VALUES (?,?,?,?,?,?,?)",
                            (aluno_id, disc_nome, ano_letivo, cif_raw, ex1, ex2, cfd_raw)
                        )
                    importados += 1

            db.commit()
            msg = f"✓ {importados} registos importados do AluDisc."
            if sem_bi_txt:
                msg += f" {len(sem_bi_txt)} BI(s) não encontrado(s)."
            flash(msg, "success" if not sem_bi_txt else "warning")
            return redirect(url_for("importar_aludisc"))

        db = get_db()

        # Construir mapa BI → {ano_letivo → aluno_id}
        bi_map = {}
        for a in db.execute("SELECT id, bi, ano_letivo FROM alunos WHERE bi IS NOT NULL AND bi != ''").fetchall():
            bi_map.setdefault(a["bi"], {})[a["ano_letivo"]] = a["id"]

        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb.active

            # Detectar linha de cabeçalho
            header_row = None
            for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
                if row[0] and "BI" in str(row[0]).upper():
                    header_row = i
                    break
            if not header_row:
                flash("Cabeçalho não encontrado no ficheiro.", "danger")
                return redirect(url_for("importar_aludisc"))

            import re as _re2
            importados = 0
            sem_bi = set()
            codigos_desconhecidos = set()

            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                bi_raw  = str(row[0]).strip() if row[0] else ""
                cod     = str(row[1]).strip() if row[1] else ""
                ano_raw = str(row[2]).strip() if row[2] else ""
                cif_raw = row[6]
                ex1_raw = row[8]
                ex2_raw = row[9]
                cfd_raw = row[10]

                if not bi_raw or not cod:
                    continue

                # Normalizar BI: primeiros 8 dígitos (ignora dígitos de controlo)
                bi = _re2.sub(r'[^0-9]', '', bi_raw)[:8]

                # Determinar ano letivo (ex: 2025 → "2024/2025")
                try:
                    ano_n = int(float(ano_raw))
                    ano_letivo = f"{ano_n-1}/{ano_n}"
                except Exception:
                    continue

                # Nome da disciplina
                disc_nome = MAPA_CODIGOS.get(cod)
                if not disc_nome:
                    codigos_desconhecidos.add(cod)
                    disc_nome = cod  # usar código como nome

                # Converter notas
                def parse_n(v):
                    if v is None: return None
                    try: return float(str(v).strip())
                    except: return None

                cif = parse_n(cif_raw)
                ex1 = parse_n(ex1_raw)
                ex2 = parse_n(ex2_raw)
                cfd = parse_n(cfd_raw)

                # Converter exame de escala 0-200 para 0-20
                if ex1 is not None and ex1 > 20: ex1 = round(ex1 / 10, 1)
                if ex2 is not None and ex2 > 20: ex2 = round(ex2 / 10, 1)

                # Encontrar aluno pelo BI — usar o mais recente disponível
                aluno_id = None
                if bi in bi_map:
                    anos_disponiveis = bi_map[bi]
                    aluno_id = (anos_disponiveis.get(ano_letivo)
                                or anos_disponiveis.get(max(anos_disponiveis.keys())))

                if aluno_id is None:
                    sem_bi.add(bi)
                    continue

                # Upsert em notas_finais
                ex = db.execute(
                    "SELECT id FROM notas_finais WHERE aluno_id=? AND disciplina=? AND ano_letivo=?",
                    (aluno_id, disc_nome, ano_letivo)
                ).fetchone()
                if ex:
                    db.execute(
                        "UPDATE notas_finais SET cif=?, exame_f1=?, exame_f2=?, cfd=? WHERE id=?",
                        (cif, ex1, ex2, cfd, ex["id"])
                    )
                else:
                    db.execute(
                        "INSERT INTO notas_finais (aluno_id, disciplina, ano_letivo, cif, exame_f1, exame_f2, cfd) VALUES (?,?,?,?,?,?,?)",
                        (aluno_id, disc_nome, ano_letivo, cif, ex1, ex2, cfd)
                    )
                importados += 1

            db.commit()
            msg = f"✓ {importados} registos importados."
            if sem_bi:
                msg += f" {len(sem_bi)} BI(s) não encontrados (alunos sem BI registado)."
            if codigos_desconhecidos:
                msg += f" Códigos desconhecidos: {', '.join(sorted(codigos_desconhecidos))}."
            flash(msg, "success" if not sem_bi else "warning")

        except Exception as e:
            flash(f"Erro: {e}", "danger")

        return redirect(url_for("importar_aludisc"))

    # Contagem actual
    try:
        n = db.execute("SELECT COUNT(*) FROM notas_finais").fetchone()[0]
    except Exception:
        n = 0
    return render_template("importar_aludisc.html", n_registos=n)


# ─── Importar inscrições de exame (AluExame.txt) ──────────────────────────────

@app.route("/admin/importar-aluexame", methods=["GET", "POST"])
@login_required
@admin_required
def importar_aluexame():
    db = get_db()
    ano = ano_letivo_atual(db)

    if request.method == "POST":
        f = request.files.get("ficheiro")
        if not f or not f.filename.endswith(".txt"):
            flash("Por favor carregue um ficheiro .txt.", "danger")
            return redirect(url_for("importar_aluexame"))

        path = os.path.join(UPLOAD_FOLDER, secure_filename(f.filename))
        f.save(path)

        # Construir mapa BI → {ano_letivo → aluno_id}
        bi_map = {}
        for a in db.execute("SELECT id, bi, ano_letivo FROM alunos WHERE bi IS NOT NULL AND bi != ''").fetchall():
            bi_map.setdefault(a["bi"], {})[a["ano_letivo"]] = a["id"]

        importados = 0
        sem_bi = set()

        with open(path, encoding="latin-1") as fh:
            for line in fh:
                line = line.rstrip("\n")
                if len(line) < 19: continue
                bi        = line[1:9].strip()
                cod_exame = line[10:13].strip()
                interno   = line[15:16].strip() or "N"
                aprovacao = line[16:17].strip() or "N"
                melhoria  = line[17:18].strip() or "N"

                if not bi or not cod_exame: continue

                disc_nome = MAPA_COD_EXAME.get(cod_exame)
                if not disc_nome: continue  # código desconhecido

                # Encontrar aluno pelo BI
                aluno_id = None
                if bi in bi_map:
                    anos = bi_map[bi]
                    aluno_id = anos.get(ano) or anos.get(max(anos.keys()))
                if aluno_id is None:
                    sem_bi.add(bi); continue

                ex = db.execute(
                    "SELECT id FROM inscricoes_exame WHERE aluno_id=? AND cod_exame=? AND ano_letivo=?",
                    (aluno_id, cod_exame, ano)
                ).fetchone()
                if ex:
                    db.execute(
                        "UPDATE inscricoes_exame SET disciplina=?, interno=?, aprovacao=?, melhoria=? WHERE id=?",
                        (disc_nome, interno, aprovacao, melhoria, ex["id"])
                    )
                else:
                    db.execute(
                        "INSERT INTO inscricoes_exame (aluno_id, disciplina, cod_exame, interno, aprovacao, melhoria, ano_letivo) VALUES (?,?,?,?,?,?,?)",
                        (aluno_id, disc_nome, cod_exame, interno, aprovacao, melhoria, ano)
                    )
                importados += 1

        db.commit()
        msg = f"✓ {importados} inscrições importadas."
        if sem_bi:
            msg += f" {len(sem_bi)} BI(s) não encontrados."
        flash(msg, "success" if not sem_bi else "warning")
        n = db.execute("SELECT COUNT(*) FROM inscricoes_exame WHERE ano_letivo=?", (ano,)).fetchone()[0]
        return render_template("importar_aluexame.html", n_registos=n, ano=ano,
                               sem_bi=sorted(sem_bi))

    n = db.execute("SELECT COUNT(*) FROM inscricoes_exame WHERE ano_letivo=?", (ano,)).fetchone()[0]
    return render_template("importar_aluexame.html", n_registos=n, ano=ano, sem_bi=[])


# ─── Importar classificações de exame (AluExame.txt) ─────────────────────────

@app.route("/admin/importar-notas-exame", methods=["GET", "POST"])
@login_required
@admin_required
def importar_notas_exame():
    """Importa classificações de exame do ficheiro AluExame.txt.
    Estrutura (posições 1-based, ficheiro com espaço inicial na posição 0):
      1-8   BI          9  Tipo BI
      10-12 Cód. exame  13 Fase (1/2)
      14 Pauta  15 Aprovação  16 Melhoria  17 Ingresso
      18-21 Cód. disciplina
      22-24 Classificação exame (0-200; -1 = faltou)
      25-27 Comp. escrita  28-30 Comp. oral
    """
    db = get_db()
    ano = ano_letivo_atual(db)

    if request.method == "POST":
        f = request.files.get("ficheiro")
        if not f or not f.filename.endswith(".txt"):
            flash("Por favor carregue um ficheiro .txt.", "danger")
            return redirect(url_for("importar_notas_exame"))

        path = os.path.join(UPLOAD_FOLDER, secure_filename(f.filename))
        f.save(path)

        # Construir mapa BI → {ano_letivo → aluno_id}
        bi_map = {}
        for a in db.execute("SELECT id, bi, ano_letivo FROM alunos WHERE bi IS NOT NULL AND bi != ''").fetchall():
            bi_map.setdefault(a["bi"], {})[a["ano_letivo"]] = a["id"]

        importados = 0
        sem_bi = set()
        faltaram = 0

        def _pi(s):
            s = s.strip()
            if not s or s == "-1": return None
            try: return int(s)
            except: return None

        with open(path, encoding="latin-1") as fh:
            for line in fh:
                line = line.rstrip("\n")
                if len(line) < 25: continue

                bi        = line[1:9].strip()
                cod_exame = line[10:13].strip()
                fase      = line[13].strip()
                nota_raw  = _pi(line[22:25])

                if not bi or not cod_exame or fase not in ("1", "2"): continue

                # "-1" = faltou → ignorar
                raw_field = line[22:25].strip()
                if raw_field == "-1":
                    faltaram += 1
                    continue

                if nota_raw is None: continue

                disc_nome = MAPA_COD_EXAME.get(cod_exame)
                if not disc_nome: continue

                # Converter 0-200 → 0-20
                nota = round(nota_raw / 10, 1)

                # Encontrar aluno pelo BI
                aluno_id = None
                if bi in bi_map:
                    anos = bi_map[bi]
                    aluno_id = anos.get(ano) or anos.get(max(anos.keys()))
                if aluno_id is None:
                    sem_bi.add(bi); continue

                campo_exame = "exame_f1" if fase == "1" else "exame_f2"

                ex = db.execute(
                    "SELECT id FROM notas_finais WHERE aluno_id=? AND disciplina=? AND ano_letivo=?",
                    (aluno_id, disc_nome, ano)
                ).fetchone()
                if ex:
                    db.execute(
                        f"UPDATE notas_finais SET {campo_exame}=? WHERE id=?",
                        (nota, ex["id"])
                    )
                else:
                    f1 = nota if fase == "1" else None
                    f2 = nota if fase == "2" else None
                    db.execute(
                        "INSERT INTO notas_finais (aluno_id, disciplina, ano_letivo, exame_f1, exame_f2) VALUES (?,?,?,?,?)",
                        (aluno_id, disc_nome, ano, f1, f2)
                    )
                importados += 1

        db.commit()
        msg = f"✓ {importados} classificações importadas."
        if faltaram:
            msg += f" {faltaram} registo(s) ignorado(s) (aluno faltou)."
        if sem_bi:
            msg += f" {len(sem_bi)} BI(s) não encontrado(s)."
        flash(msg, "success" if not sem_bi else "warning")
        n = db.execute(
            "SELECT COUNT(*) FROM notas_finais WHERE ano_letivo=? AND (exame_f1 IS NOT NULL OR exame_f2 IS NOT NULL)",
            (ano,)
        ).fetchone()[0]
        return render_template("importar_notas_exame.html", n_registos=n, ano=ano, sem_bi=sorted(sem_bi))

    n = db.execute(
        "SELECT COUNT(*) FROM notas_finais WHERE ano_letivo=? AND (exame_f1 IS NOT NULL OR exame_f2 IS NOT NULL)",
        (ano,)
    ).fetchone()[0]
    return render_template("importar_notas_exame.html", n_registos=n, ano=ano, sem_bi=[])


@app.route("/admin/importar-notas", methods=["GET", "POST"])
@login_required
@admin_required
def importar_notas_web():
    """Importa ficheiros de Avaliação Contínua (formato Grelha)."""
    if request.method == "POST":
        ficheiros = request.files.getlist("ficheiros")
        semestre  = int(request.form.get("semestre", 1))
        ano       = request.form.get("ano_letivo", "2025/2026")

        if not ficheiros or all(f.filename == "" for f in ficheiros):
            flash("Selecione pelo menos um ficheiro.", "danger")
            return redirect(url_for("importar_notas_web"))

        import re as _re, unicodedata as _uc

        def _normalizar(s):
            s = _uc.normalize("NFKD", str(s or ""))
            s = s.encode("ascii", "ignore").decode()
            return _re.sub(r"\s+", " ", s).strip().lower()

        def _parse_nota(v):
            if v is None: return None
            s = str(v).strip()
            if s in ("-", "", "NP", "NP.", "NE", "NA", "—"): return None
            try: return float(s)
            except: return None

        def _extrair_turma(val):
            s = str(val).strip()
            return s.split(" - ", 1)[1].strip() if " - " in s else s

        # Mapeamento de nomes abreviados (formato legado) → nomes completos
        MAPA_DISC = {
            "PORT.": "Português", "FILO.": "Filosofia",
            "ED.FÍSICA": "Educação Física", "RELIGIÃO": "Religião",
            "LE I-ING.": "Líng. Estrang. I - Inglês",
            "MAT.A": "Matemática A", "MAT.B": "Matemática Geral",
            "MAT.GERAL": "Matemática Geral",
            "BIO.GEO.": "Biologia e Geologia",
            "FÍS.QUÍM.A": "Física e Química A", "FIS.QUIM.A": "Física e Química A",
            "GEOG.A": "Geografia A", "HIST.A": "História A",
            "HIST. B": "História Geral", "HIST.B": "História Geral",
            "DES.A": "Desenho A", "DES.GERAL": "Desenho Geral",
            "ECON.A": "Economia A", "ECON.C": "Economia C",
            "GEO.DESC.A": "Geometria Descritiva A",
            "MACS": "Matemática Aplicada Ciências Sociais",
            "BIO.": "Biologia", "FIS.": "Física", "QUIM.": "Química",
            "PSIC.B": "Psicologia B", "AI.B": "Aplicações Informáticas B",
            "C.POL.": "Ciência Política", "FILO.A": "Filosofia A",
        }

        def _nome_disc(abrev):
            """Converte abreviatura para nome completo, se existir mapeamento."""
            return MAPA_DISC.get(abrev.strip(), abrev.strip())

        def _turma_legado(val):
            """'23/24 - 10º ANO A1' → '10A1'"""
            m = _re.search(r"(\d+)º ANO ([A-Z]\d+)", str(val or ""))
            if m:
                return m.group(1) + m.group(2)  # ex: "10A1"
            return _extrair_turma(val)

        def _detectar_formato_flat(ws):
            """Detecta se o ficheiro é formato flat (uma linha por aluno×disciplina)."""
            headers = [str(c.value or "").strip().lower() for c in ws[1]]
            return any("disciplina" in h for h in headers) and any("nome do aluno" in h or "nome aluno" in h for h in headers)

        def _importar_formato_flat(ws, db, semestre, ano, alunos_cache, alunos_todos):
            """Importa ficheiro flat: uma linha por aluno×disciplina."""
            headers = [str(c.value or "").strip() for c in ws[1]]
            h = [x.lower() for x in headers]

            def ci(names):
                for n in names:
                    for i, hh in enumerate(h):
                        if n.lower() in hh: return i
                return None

            idx_num   = ci(["nº processo", "numero processo", "n.º processo", "aluno (nº"])
            idx_nome  = ci(["nome do aluno", "nome aluno"])
            idx_turma = ci(["turma"])
            idx_disc  = ci(["disciplina"])
            idx_1s    = ci(["nota 1º semestre", "nota 1o semestre", "1º semestre"])
            idx_2s    = ci(["nota 2º semestre", "nota 2o semestre", "2º semestre"])
            idx_ano   = ci(["ano letivo"])
            idx_bi    = ci(["documento de identificação", "bi", "doc. identificação", "documento identificacao"])

            if idx_nome is None or idx_disc is None:
                return 0, 0, set()

            criados = 0
            total_notas = 0
            nao_enc = set()

            for row in ws.iter_rows(min_row=2, values_only=True):
                nome_val = row[idx_nome] if idx_nome is not None else None
                if not nome_val or not str(nome_val).strip():
                    continue

                nome_str = str(nome_val).strip()
                nome_n   = _normalizar(nome_str)
                num_val  = str(row[idx_num]).strip() if idx_num is not None and row[idx_num] else ""
                turma_raw = row[idx_turma] if idx_turma is not None else ""
                turma_val = _turma_legado(turma_raw)
                ano_val   = str(row[idx_ano]).strip() if idx_ano is not None and row[idx_ano] else ano
                disc_raw  = str(row[idx_disc]).strip() if row[idx_disc] else ""
                disc_nome = _nome_disc(disc_raw)

                # Encontrar ou criar aluno para este ano letivo
                aluno_id = alunos_cache.get((nome_n, turma_val))
                if aluno_id is None:
                    numero_aluno = num_val
                    if not numero_aluno and nome_n in alunos_todos:
                        numero_aluno = alunos_todos[nome_n][0]["numero"] or ""
                    try:
                        cur = db.execute(
                            "INSERT OR IGNORE INTO alunos (numero, nome, turma, ano_letivo) VALUES (?,?,?,?)",
                            (numero_aluno, nome_str, turma_val, ano_val)
                        )
                        if cur.rowcount > 0:
                            aluno_id = cur.lastrowid
                            criados += 1
                        else:
                            r = db.execute(
                                "SELECT id FROM alunos WHERE nome=? AND turma=? AND ano_letivo=?",
                                (nome_str, turma_val, ano_val)
                            ).fetchone()
                            if r: aluno_id = r["id"]
                        if aluno_id:
                            alunos_cache[(nome_n, turma_val)] = aluno_id
                            alunos_todos.setdefault(nome_n, []).append({"id": aluno_id, "numero": numero_aluno})
                    except Exception:
                        pass

                if aluno_id is None:
                    nao_enc.add(f"{nome_str} ({turma_val})")
                    continue

                # Guardar BI (normalizado: primeiros 8 dígitos — sempre actualizar)
                if idx_bi is not None and row[idx_bi]:
                    bi_digits = _re.sub(r'[^0-9]', '', str(row[idx_bi]))
                    bi_val = bi_digits[:8]
                    if bi_val:
                        try:
                            db.execute("UPDATE alunos SET bi=? WHERE id=?", (bi_val, aluno_id))
                        except Exception:
                            pass

                # Importar nota 1º semestre
                n1 = _parse_nota(row[idx_1s] if idx_1s is not None else None)
                if n1 is not None:
                    ex = db.execute("SELECT id FROM notas WHERE aluno_id=? AND disciplina=? AND periodo=1",
                                    (aluno_id, disc_nome)).fetchone()
                    if ex: db.execute("UPDATE notas SET nota=? WHERE id=?", (n1, ex["id"]))
                    else:  db.execute("INSERT INTO notas (aluno_id, disciplina, periodo, nota) VALUES (?,?,1,?)",
                                      (aluno_id, disc_nome, n1))
                    total_notas += 1

                # Importar nota 2º semestre
                n2 = _parse_nota(row[idx_2s] if idx_2s is not None else None)
                if n2 is not None:
                    ex = db.execute("SELECT id FROM notas WHERE aluno_id=? AND disciplina=? AND periodo=2",
                                    (aluno_id, disc_nome)).fetchone()
                    if ex: db.execute("UPDATE notas SET nota=? WHERE id=?", (n2, ex["id"]))
                    else:  db.execute("INSERT INTO notas (aluno_id, disciplina, periodo, nota) VALUES (?,?,2,?)",
                                      (aluno_id, disc_nome, n2))
                    total_notas += 1

            return total_notas, criados, nao_enc

        db = get_db()

        # Cache alunos para o ano pretendido
        alunos_cache = {}
        for a in db.execute("SELECT id, nome, turma FROM alunos WHERE ano_letivo=?", (ano,)).fetchall():
            alunos_cache[(_normalizar(a["nome"]), a["turma"])] = a["id"]

        # Cache por nome normalizado em TODOS os anos (para criar registos de anos anteriores)
        alunos_todos = {}
        for a in db.execute("SELECT id, nome, numero FROM alunos").fetchall():
            alunos_todos.setdefault(_normalizar(a["nome"]), []).append({"id": a["id"], "numero": a["numero"]})

        total_notas   = 0
        criados_auto  = 0
        nao_enc       = set()

        for f in ficheiros:
            if not f.filename.endswith((".xlsx", ".xls")):
                continue
            path = os.path.join(UPLOAD_FOLDER, secure_filename(f.filename))
            f.save(path)

            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb.active

            # ── Detecção automática do formato ─────────────────────────────
            if _detectar_formato_flat(ws):
                # Formato flat/legado (uma linha por aluno×disciplina)
                n, c, ne = _importar_formato_flat(ws, db, semestre, ano, alunos_cache, alunos_todos)
                total_notas  += n
                criados_auto += c
                nao_enc      |= ne
                continue

            # ── Formato pauta oficial (PORT./FILO./etc. como cabeçalhos) ───
            _turma_doc, _ano_doc, _sem_doc, _alunos_doc = parse_pauta_excel(ws)
            if _alunos_doc:
                _ano_usar = _ano_doc or ano
                _sem_usar = _sem_doc or semestre
                # Nível curricular da turma detectada na pauta
                _nivel_pauta2 = int((_turma_doc or "")[:2]) if (_turma_doc or "")[:2].isdigit() else 0
                for _al in _alunos_doc:
                    _aid = alunos_cache.get((_normalizar(_al["nome"]), None))
                    if _aid is None:
                        r = db.execute(
                            "SELECT id FROM alunos WHERE numero=? AND ano_letivo=?",
                            (_al["numero"], _ano_usar)
                        ).fetchone()
                        if r: _aid = r["id"]
                    if _aid is None:
                        for _a_db in db.execute(
                            "SELECT id, nome FROM alunos WHERE ano_letivo=?", (_ano_usar,)
                        ).fetchall():
                            if nome_match(_al["nome"], _a_db["nome"]):
                                _aid = _a_db["id"]; break
                    if _aid is None:
                        nao_enc.add(_al["nome"]); continue
                    for _disc_abrev, _campos in _al["notas"].items():
                        _disc_nome = MAPA_DISC_GLOBAL.get(_disc_abrev, _disc_abrev)
                        if _nivel_pauta2 == 11 and _disc_nome == "Desenho Geral":
                            _disc_nome = "Desenho A"
                        _cf = _campos.get("CF")
                        if _cf is not None:
                            # Apagar registos de disciplinas equivalentes (mesma família) para o mesmo período
                            _fam_membros = membros_familia(_disc_nome)
                            for _membro in _fam_membros:
                                if _membro != _disc_nome:
                                    db.execute(
                                        "DELETE FROM notas WHERE aluno_id=? AND disciplina=? AND periodo=?",
                                        (_aid, _membro, _sem_usar)
                                    )
                            _ex = db.execute(
                                "SELECT id FROM notas WHERE aluno_id=? AND disciplina=? AND periodo=?",
                                (_aid, _disc_nome, _sem_usar)
                            ).fetchone()
                            if _ex:
                                db.execute("UPDATE notas SET nota=? WHERE id=?", (_cf, _ex["id"]))
                            else:
                                db.execute(
                                    "INSERT INTO notas (aluno_id, disciplina, periodo, nota) VALUES (?,?,?,?)",
                                    (_aid, _disc_nome, _sem_usar, _cf)
                                )
                            total_notas += 1
                db.commit()
                continue

            # ── Formato grelha (largo, com NP.) ────────────────────────────

            rows = list(ws.iter_rows(values_only=True))

            header_idx = None
            for i, row in enumerate(rows):
                if row[0] and "Nome" in str(row[0]) and "Turma" in str(row[0]):
                    header_idx = i; break
            if header_idx is None:
                flash(f"Estrutura não reconhecida: {f.filename}", "warning")
                continue

            disc_row = rows[header_idx - 2]
            np_row   = rows[header_idx]
            disc_cols_sorted = sorted([i for i, v in enumerate(disc_row)
                                       if v and str(v).strip()
                                       and "Nome" not in str(v)
                                       and "Número" not in str(v)])
            disc_np_map = {}
            for idx, dc in enumerate(disc_cols_sorted):
                end = disc_cols_sorted[idx+1] if idx+1 < len(disc_cols_sorted) else len(np_row)
                dname = str(disc_row[dc]).strip()
                for c in range(dc, min(end, len(np_row))):
                    if np_row[c] and str(np_row[c]).strip() == "NP.":
                        disc_np_map[dname] = c; break

            # Detectar coluna do número na grelha (col 4 = Número na turma, não útil;
            # col 1 do header pode ter "N.º Matrícula" — já tratado em parse_pauta_excel)
            col_num_grelha = next(
                (ci for ci, v in enumerate(rows[header_idx]) if v and "N.º" in str(v) and "Matrícula" in str(v)), None
            )

            turma_atual = None
            for row in rows[header_idx + 2:]:
                if row[0] and str(row[0]).strip():
                    turma_atual = _extrair_turma(row[0])
                nome_val = row[7] if len(row) > 7 else None
                if not nome_val or not str(nome_val).strip(): continue
                nome_n   = _normalizar(str(nome_val))
                nome_str = str(nome_val).strip()
                # Número de matrícula da grelha (col 1 nas pautas, col 4 nas grelhas de avaliação)
                num_grelha = str(row[col_num_grelha]).strip() if col_num_grelha and col_num_grelha < len(row) and row[col_num_grelha] else \
                             str(row[1]).strip() if len(row) > 1 and row[1] and str(row[1]).strip().isdigit() else ""

                # 1. Por número de matrícula (mais fiável)
                aluno_id = None
                if num_grelha:
                    r = db.execute(
                        "SELECT id FROM alunos WHERE numero=? AND ano_letivo=?", (num_grelha, ano)
                    ).fetchone()
                    if r: aluno_id = r["id"]

                # 2. Por nome exacto
                if aluno_id is None:
                    aluno_id = alunos_cache.get((nome_n, turma_atual))

                # 3. Por nome fuzzy (iniciais)
                if aluno_id is None:
                    cands = db.execute(
                        "SELECT id, nome FROM alunos WHERE ano_letivo=? AND turma=?", (ano, turma_atual or "")
                    ).fetchall()
                    for c in cands:
                        if nome_match(nome_str, c["nome"]):
                            aluno_id = c["id"]
                            break

                # 4. Se não encontrado, criar registo
                if aluno_id is None:
                    numero_aluno = num_grelha
                    if not numero_aluno and nome_n in alunos_todos:
                        numero_aluno = alunos_todos[nome_n][0]["numero"] or ""

                    # Criar registo do aluno para este ano letivo
                    try:
                        cur = db.execute(
                            "INSERT OR IGNORE INTO alunos (numero, nome, turma, ano_letivo) VALUES (?,?,?,?)",
                            (numero_aluno, nome_str, turma_atual or "", ano)
                        )
                        if cur.rowcount > 0:
                            aluno_id = cur.lastrowid
                            alunos_cache[(nome_n, turma_atual)] = aluno_id
                            alunos_todos.setdefault(nome_n, []).append({"id": aluno_id, "numero": numero_aluno})
                            criados_auto += 1
                        else:
                            # Já existe — buscar id
                            r = db.execute(
                                "SELECT id FROM alunos WHERE nome=? AND turma=? AND ano_letivo=?",
                                (nome_str, turma_atual or "", ano)
                            ).fetchone()
                            if r:
                                aluno_id = r["id"]
                                alunos_cache[(nome_n, turma_atual)] = aluno_id
                    except Exception:
                        pass

                if aluno_id is None:
                    nao_enc.add(f"{nome_val} ({turma_atual})")
                    continue

                for disc, col_np in disc_np_map.items():
                    nota = _parse_nota(row[col_np] if col_np < len(row) else None)
                    if nota is None: continue
                    ex = db.execute(
                        "SELECT id FROM notas WHERE aluno_id=? AND disciplina=? AND periodo=?",
                        (aluno_id, disc, semestre)
                    ).fetchone()
                    if ex:
                        db.execute("UPDATE notas SET nota=? WHERE id=?", (nota, ex["id"]))
                    else:
                        db.execute(
                            "INSERT INTO notas (aluno_id, disciplina, periodo, nota) VALUES (?,?,?,?)",
                            (aluno_id, disc, semestre, nota)
                        )
                    total_notas += 1

        db.commit()
        msg = f"Importação concluída: {total_notas} notas carregadas."
        if criados_auto:
            msg += f" {criados_auto} aluno(s) criado(s) automaticamente para {ano}."
        if nao_enc:
            msg += f" {len(nao_enc)} aluno(s) não encontrado(s)."
        flash(msg, "success" if not nao_enc else "warning")
        return redirect(url_for("importar_notas_web"))

    return render_template("importar_notas.html")


# ─── Run ───────────────────────────────────────────────────────────────────────

# Garantir que a BD existe ao arrancar (gunicorn ou python app.py)
try:
    init_db()
except Exception as _e:
    print(f"[AVISO] init_db: {_e}")

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
