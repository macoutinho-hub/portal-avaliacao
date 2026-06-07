"""
Importa os dados de Avaliação de Projeto a partir de UM ÚNICO ficheiro
consolidado ('Avaliacao_Projeto_11ano_consolidado.xlsx'), já preparado com
toda a informação dos vários Excel originais (avaliações, grupos, temas,
questões orientadoras, orientadores e membros — já feito o "matching" de nomes).

Sheets esperadas:
  - Avaliacoes: Numero, Nome, Turma, Media Workshops, Desempenho Aula,
    Desempenho Aula (nivel), Apresentacao Oral QO, Poster, Questionario,
    Artigo, Media Componentes, Apresentacao Final, Avaliacao Produto Final,
    Avaliacao Final, Observacao
  - Grupos: Sala, Numero Grupo, Tema, Questao Orientadora, Estado Aprovacao,
    Professor Orientador, Membros (nome - turma)   [coluna informativa, não usada na importação]
  - Membros: Numero Aluno, Nome Aluno, Turma, Sala, Numero Grupo

Uso:
    python importar_projeto_consolidado.py [Avaliacao_Projeto_11ano_consolidado.xlsx]
"""

import sqlite3
import sys
import os
import openpyxl

DATABASE   = "portal.db"
ANO_LETIVO = "2025/2026"
NIVEL      = "11º ano"
SEMESTRE   = 1

FICHEIRO_PADRAO = "Avaliacao_Projeto_11ano_consolidado.xlsx"


def parse_num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).strip().replace(",", "."))
    except ValueError:
        return None


def _aluno_id(db, numero, nome, turma):
    numero = str(numero).strip()
    aluno = db.execute("SELECT id FROM alunos WHERE numero=? AND ano_letivo=?",
                       (numero, ANO_LETIVO)).fetchone()
    if not aluno:
        aluno = db.execute(
            "SELECT id FROM alunos WHERE turma=? AND ano_letivo=? AND lower(nome)=lower(?)",
            (str(turma).strip(), ANO_LETIVO, str(nome).strip())).fetchone()
    return aluno["id"] if aluno else None


def importar_avaliacoes(db, ws):
    cabecalho = [c.value for c in ws[1]]
    n_ok = n_falhas = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        numero, nome, turma = row[0], row[1], row[2]
        aluno_id = _aluno_id(db, numero, nome, turma)
        if aluno_id is None:
            print(f"  ! aluno não encontrado: {numero} - {nome} ({turma})")
            n_falhas += 1
            continue

        (media_workshops, desempenho_aula, desempenho_aula_nivel, apresentacao_oral_qo,
         poster, questionario, artigo, media_componentes, apresentacao_final,
         avaliacao_produto_final, avaliacao_final, observacao) = row[3:15]

        db.execute("""
            INSERT INTO projeto_avaliacao
                (aluno_id, ano_letivo, semestre, media_workshops, desempenho_aula,
                 desempenho_aula_nivel, apresentacao_oral_qo, poster, questionario,
                 artigo, media_componentes, apresentacao_final, avaliacao_produto_final,
                 avaliacao_final, observacao)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            ON CONFLICT(aluno_id, ano_letivo, semestre) DO UPDATE SET
                media_workshops=excluded.media_workshops,
                desempenho_aula=excluded.desempenho_aula,
                desempenho_aula_nivel=excluded.desempenho_aula_nivel,
                apresentacao_oral_qo=excluded.apresentacao_oral_qo,
                poster=excluded.poster,
                questionario=excluded.questionario,
                artigo=excluded.artigo,
                media_componentes=excluded.media_componentes,
                apresentacao_final=excluded.apresentacao_final,
                avaliacao_produto_final=excluded.avaliacao_produto_final,
                avaliacao_final=excluded.avaliacao_final,
                observacao=excluded.observacao
        """, (aluno_id, ANO_LETIVO, SEMESTRE, parse_num(media_workshops), parse_num(desempenho_aula),
              desempenho_aula_nivel, parse_num(apresentacao_oral_qo), parse_num(poster),
              parse_num(questionario), parse_num(artigo), parse_num(media_componentes),
              parse_num(apresentacao_final), parse_num(avaliacao_produto_final),
              parse_num(avaliacao_final), observacao))
        n_ok += 1
    db.commit()
    print(f"  → {n_ok} avaliações importadas/atualizadas, {n_falhas} alunos não encontrados")


def importar_grupos(db, ws_grupos, ws_membros):
    # limpar estado anterior
    db.execute("""DELETE FROM projeto_grupo_membros WHERE grupo_id IN
                  (SELECT id FROM projeto_grupos WHERE ano_letivo=? AND nivel=?)""", (ANO_LETIVO, NIVEL))
    db.execute("DELETE FROM projeto_grupos WHERE ano_letivo=? AND nivel=?", (ANO_LETIVO, NIVEL))
    db.commit()

    grupo_id_por_chave = {}
    n_grupos = 0
    for row in ws_grupos.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        sala, numero, tema, questao, estado, orientador = row[0], row[1], row[2], row[3], row[4], row[5]
        cur = db.execute("""
            INSERT INTO projeto_grupos (ano_letivo, nivel, sala, numero, tema,
                                        questao_orientadora, estado_aprovacao, professor_orientador)
            VALUES (?,?,?,?,?,?,?,?)
            ON CONFLICT(ano_letivo, nivel, sala, numero) DO UPDATE SET
                tema=excluded.tema, questao_orientadora=excluded.questao_orientadora,
                estado_aprovacao=excluded.estado_aprovacao, professor_orientador=excluded.professor_orientador
        """, (ANO_LETIVO, NIVEL, sala, numero, tema, questao, estado, orientador))
        grupo_id = cur.lastrowid or db.execute(
            "SELECT id FROM projeto_grupos WHERE ano_letivo=? AND nivel=? AND sala=? AND numero=?",
            (ANO_LETIVO, NIVEL, sala, numero)).fetchone()["id"]
        grupo_id_por_chave[(sala, numero)] = grupo_id
        n_grupos += 1
    db.commit()

    n_membros = n_falhas = 0
    for row in ws_membros.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        numero_aluno, nome, turma, sala, numero_grupo = row[0], row[1], row[2], row[3], row[4]
        grupo_id = grupo_id_por_chave.get((sala, numero_grupo))
        if grupo_id is None:
            continue
        aluno_id = _aluno_id(db, numero_aluno, nome, turma)
        if aluno_id is None:
            print(f"  ! membro não encontrado: {numero_aluno} - {nome} ({turma})")
            n_falhas += 1
            continue
        db.execute("INSERT OR IGNORE INTO projeto_grupo_membros (grupo_id, aluno_id) VALUES (?,?)",
                   (grupo_id, aluno_id))
        n_membros += 1
    db.commit()
    print(f"  → {n_grupos} grupos, {n_membros} membros associados, {n_falhas} membros não encontrados")


def main():
    ficheiro = sys.argv[1] if len(sys.argv) > 1 else FICHEIRO_PADRAO
    if not os.path.exists(ficheiro):
        print(f"! ficheiro não encontrado: {ficheiro}")
        return

    db = sqlite3.connect(DATABASE)
    db.row_factory = sqlite3.Row
    db.execute("PRAGMA foreign_keys = ON")

    wb = openpyxl.load_workbook(ficheiro, data_only=True)

    print(f"A importar avaliações de '{ficheiro}' (sheet 'Avaliacoes')…")
    importar_avaliacoes(db, wb["Avaliacoes"])

    if "Grupos" in wb.sheetnames and "Membros" in wb.sheetnames:
        print("A importar grupos/temas/orientadores/membros…")
        importar_grupos(db, wb["Grupos"], wb["Membros"])

    db.close()
    print("Concluído.")


if __name__ == "__main__":
    main()
