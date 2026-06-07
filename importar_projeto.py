"""
Importa os dados de Avaliação de Projeto do 11º ano para a BD do portal:
  - projeto_avaliacao        (componentes e nota final por aluno, sheet AVAL_PROJ)
  - projeto_grupos           (tema, questão orientadora, professor orientador)
  - projeto_grupo_membros    (alunos de cada grupo — grupos são transversais a turmas)

Uso:
    python importar_projeto.py "2526_Avaliação de Projeto 11.º Ano - Cópia.xlsx" \
                               "Grupos_Temas_professores_orientadores.xlsx" \
                               "Grupos e questões orientadoras.xlsx"

Por omissão procura estes três ficheiros na pasta atual (nomes podem ser
ajustados nas constantes abaixo).
"""

import sqlite3
import sys
import os
import re
import unicodedata
import openpyxl

DATABASE   = "portal.db"
ANO_LETIVO = "2025/2026"
NIVEL      = "11º ano"
SEMESTRE   = 1

F_AVALIACAO = "2526_Avaliação de Projeto 11.º Ano - Cópia.xlsx"
F_GRUPOS    = "Grupos_Temas_professores_orientadores.xlsx"
F_QUESTOES  = "Grupos e questões orientadoras.xlsx"


# ─── Utilitários ──────────────────────────────────────────────────────────────

def normalizar(s):
    """Remove acentos, espaços extra e converte para minúsculas (comparação fuzzy)."""
    if not s:
        return ""
    s = unicodedata.normalize("NFKD", str(s))
    s = s.encode("ascii", "ignore").decode()
    return re.sub(r"\s+", " ", s).strip().lower()


def parse_num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


TURMA_RE = re.compile(r"^1[0-2][A-E][12]$")


# ─── 1. AVAL_PROJ → projeto_avaliacao ────────────────────────────────────────

def importar_avaliacoes(db, path):
    wb = openpyxl.load_workbook(path, data_only=True)
    if "AVAL_PROJ" not in wb.sheetnames:
        print(f"  ! sheet AVAL_PROJ não encontrada em {path}")
        return 0

    ws = wb["AVAL_PROJ"]
    n_ok = n_falhas = 0

    for row in ws.iter_rows(min_row=3, values_only=True):
        turma = row[0]
        numero = row[1]
        nome = row[2]
        if not (turma and numero and nome and TURMA_RE.match(str(turma).strip())):
            continue

        numero = str(numero).strip()
        aluno = db.execute(
            "SELECT id FROM alunos WHERE numero=? AND ano_letivo=?",
            (numero, ANO_LETIVO)
        ).fetchone()
        if not aluno:
            # tentar por nome + turma (números podem ter sido reatribuídos)
            aluno = db.execute(
                "SELECT id FROM alunos WHERE turma=? AND ano_letivo=? AND lower(nome)=lower(?)",
                (str(turma).strip(), ANO_LETIVO, str(nome).strip())
            ).fetchone()
        if not aluno:
            print(f"  ! aluno não encontrado: {numero} - {nome} ({turma})")
            n_falhas += 1
            continue

        aluno_id = aluno["id"]
        media_workshops          = parse_num(row[13])
        desempenho_aula          = parse_num(row[15])
        desempenho_aula_nivel    = row[16]
        apresentacao_oral_qo     = parse_num(row[17])
        poster                   = parse_num(row[18])
        questionario             = parse_num(row[19])
        artigo                   = parse_num(row[20])
        media_componentes        = parse_num(row[21])
        apresentacao_final       = parse_num(row[22])
        avaliacao_produto_final  = parse_num(row[23])
        avaliacao_final          = parse_num(row[24])
        observacao               = row[26] if len(row) > 26 else None

        db.execute("""
            INSERT INTO projeto_avaliacao
                (aluno_id, ano_letivo, semestre, media_workshops, desempenho_aula,
                 desempenho_aula_nivel, apresentacao_oral_qo, poster, questionario,
                 artigo, media_componentes, apresentacao_final, avaliacao_produto_final,
                 avaliacao_final, observacao)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            ON CONFLICT(aluno_id, ano_letivo, semestre) DO UPDATE SET
                media_workshops=excluded.media_workshops,
                desempenho_aula=excluded.desempenho_aula,
                desempenho_aula_nivel=excluded.desempenho_aula_nivel,
                apresentacao_oral_qo=excluded.apresentacao_oral_qo,
                poster=excluded.poster,
                questionario=excluded.questionario,
                artigo=excluded.artigo,
                media_componentes=excluded.media_componentes,
                apresentacao_final=excluded.apresentacao_final,
                avaliacao_produto_final=excluded.avaliacao_produto_final,
                avaliacao_final=excluded.avaliacao_final,
                observacao=excluded.observacao
        """, (aluno_id, ANO_LETIVO, SEMESTRE, media_workshops, desempenho_aula,
              desempenho_aula_nivel, apresentacao_oral_qo, poster, questionario,
              artigo, media_componentes, apresentacao_final, avaliacao_produto_final,
              avaliacao_final, observacao))
        n_ok += 1

    db.commit()
    print(f"  → {n_ok} avaliações importadas/atualizadas, {n_falhas} alunos não encontrados")
    return n_ok


# ─── 2. Grupos / temas / orientadores ────────────────────────────────────────

def _tokens(nome):
    return [t for t in normalizar(nome).split(" ") if len(t) > 1]


def _procurar_aluno(db, nome, turma_sugerida=None):
    """Procura um aluno por nome — os ficheiros de grupos usam nomes abreviados
    ('Lara Gomes' em vez de 'Lara Sofia Fernandes Carvalho Nicho Gomes'), por
    isso comparamos por conjuntos de tokens (todas as palavras do nome
    abreviado têm de aparecer no nome completo do aluno na BD)."""
    nome_norm = normalizar(nome)
    toks_curto = set(_tokens(nome))
    if not toks_curto:
        return None

    candidatos = db.execute(
        "SELECT id, nome, turma FROM alunos WHERE ano_letivo=?", (ANO_LETIVO,)
    ).fetchall()

    exatos = [c for c in candidatos if normalizar(c["nome"]) == nome_norm]
    if len(exatos) == 1:
        return exatos[0]["id"]

    # candidatos cujo nome completo contém TODAS as palavras do nome abreviado
    subconjunto = []
    for c in candidatos:
        toks_completo = set(_tokens(c["nome"]))
        if toks_curto <= toks_completo:
            subconjunto.append((c, len(toks_completo - toks_curto)))

    if turma_sugerida:
        mesma_turma = [c for c, _ in subconjunto if normalizar(c["turma"]) == normalizar(turma_sugerida)]
        if len(mesma_turma) == 1:
            return mesma_turma[0]["id"]
        if len(mesma_turma) > 1:
            subconjunto = [(c, extra) for c, extra in subconjunto if c in mesma_turma]

    if subconjunto:
        # preferir o candidato cujo nome completo é "mais próximo" (menos palavras extra)
        subconjunto.sort(key=lambda ce: ce[1])
        if len(subconjunto) == 1 or subconjunto[0][1] < subconjunto[1][1]:
            return subconjunto[0][0]["id"]
        return None  # ambíguo

    # último recurso: sobreposição parcial de tokens (>= metade), desempatado pela turma
    parciais = []
    for c in candidatos:
        toks_completo = set(_tokens(c["nome"]))
        comuns = toks_curto & toks_completo
        if len(comuns) >= max(1, len(toks_curto) - 1) and len(comuns) >= len(toks_curto) / 2:
            parciais.append(c)
    if turma_sugerida:
        parciais_turma = [c for c in parciais if normalizar(c["turma"]) == normalizar(turma_sugerida)]
        if len(parciais_turma) == 1:
            return parciais_turma[0]["id"]
    if len(parciais) == 1:
        return parciais[0]["id"]
    return None


def _limpar_estado_grupos(db):
    db.execute("""DELETE FROM projeto_grupo_membros WHERE grupo_id IN
                  (SELECT id FROM projeto_grupos WHERE ano_letivo=? AND nivel=?)""",
               (ANO_LETIVO, NIVEL))
    db.execute("DELETE FROM projeto_grupos WHERE ano_letivo=? AND nivel=?", (ANO_LETIVO, NIVEL))
    db.commit()


def importar_grupos(db, path):
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet = wb["Folha1"] if "Folha1" in wb.sheetnames else wb[wb.sheetnames[0]]

    _limpar_estado_grupos(db)

    sala_atual = None
    grupo_atual = None      # número do grupo
    membros_pendentes = []  # [(nome, turma), ...]
    n_grupos = n_membros = n_falhas = 0

    def gravar_grupo(tema=None, orientador=None):
        nonlocal n_grupos, n_membros, n_falhas
        if grupo_atual is None or not membros_pendentes:
            return
        cur = db.execute("""
            INSERT INTO projeto_grupos (ano_letivo, nivel, sala, numero, tema, professor_orientador)
            VALUES (?,?,?,?,?,?)
            ON CONFLICT(ano_letivo, nivel, sala, numero) DO UPDATE SET
                tema=excluded.tema, professor_orientador=excluded.professor_orientador
        """, (ANO_LETIVO, NIVEL, sala_atual, grupo_atual, tema, orientador))
        grupo_id = cur.lastrowid or db.execute(
            "SELECT id FROM projeto_grupos WHERE ano_letivo=? AND nivel=? AND sala=? AND numero=?",
            (ANO_LETIVO, NIVEL, sala_atual, grupo_atual)).fetchone()["id"]
        n_grupos += 1
        for nome, turma in membros_pendentes:
            aluno_id = _procurar_aluno(db, nome, turma)
            if aluno_id is None:
                print(f"  ! membro de grupo não encontrado: '{nome}' ({turma})")
                n_falhas += 1
                continue
            db.execute("INSERT OR IGNORE INTO projeto_grupo_membros (grupo_id, aluno_id) VALUES (?,?)",
                       (grupo_id, aluno_id))
            n_membros += 1

    for row in sheet.iter_rows(values_only=True):
        c0 = row[0]
        if c0 is None:
            continue
        c0s = str(c0).strip()

        if c0s.lower().startswith("sala"):
            gravar_grupo()  # fecha grupo anterior, se existir, sem tema novo (não deveria acontecer aqui)
            sala_atual = c0s
            grupo_atual = None
            membros_pendentes = []
            continue

        m = re.match(r"grupo\s+(\d+)", c0s, re.IGNORECASE)
        if m:
            gravar_grupo()  # nunca deveria ter tema pendente aqui — tema vem na linha seguinte
            grupo_atual = int(m.group(1))
            membros_pendentes = []
            continue

        # linha de membros: pares (nome, turma, nome, turma, ...)
        if grupo_atual is not None and not membros_pendentes:
            pares = []
            for i in range(0, len(row) - 1, 2):
                nome, turma = row[i], row[i + 1]
                if nome and turma and TURMA_RE.match(str(turma).strip()):
                    pares.append((str(nome).strip(), str(turma).strip()))
            if pares:
                membros_pendentes = pares
                continue

        # linha de tema + orientador (texto longo na 1ª coluna, nome do professor mais à direita)
        if grupo_atual is not None and membros_pendentes and len(c0s) > 25:
            tema = c0s
            orientador = None
            for v in reversed(row[1:]):
                if v and isinstance(v, str) and len(v.strip()) > 2 and len(v.strip()) < 40:
                    orientador = v.strip()
                    break
            gravar_grupo(tema=tema, orientador=orientador)
            grupo_atual = None
            membros_pendentes = []
            continue

    gravar_grupo()
    db.commit()
    print(f"  → {n_grupos} grupos, {n_membros} membros associados, {n_falhas} membros não encontrados")
    return n_grupos


# ─── 3. Questões orientadoras / estado de aprovação ──────────────────────────

def importar_questoes(db, path):
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet = wb[wb.sheetnames[0]]

    grupos = db.execute(
        "SELECT id, sala, numero FROM projeto_grupos WHERE ano_letivo=? AND nivel=?",
        (ANO_LETIVO, NIVEL)
    ).fetchall()
    membros_por_grupo = {}
    for g in grupos:
        nomes = db.execute("""
            SELECT a.nome FROM projeto_grupo_membros m
            JOIN alunos a ON a.id = m.aluno_id WHERE m.grupo_id=?
        """, (g["id"],)).fetchall()
        membros_por_grupo[g["id"]] = set(normalizar(n["nome"]) for n in nomes)

    n_atualizados = 0
    cabecalho_visto = False
    for row in sheet.iter_rows(values_only=True):
        if not cabecalho_visto:
            if row and row[0] == "Número":
                cabecalho_visto = True
            continue
        if not row or row[1] is None:
            continue

        nomes_grupo = re.split(r",|;|\be\b", str(row[1]))
        nomes_grupo = set()
        for parte in re.split(r",|;", str(row[1])):
            parte = re.sub(r"\b1[0-2]º?\s*[A-E][12]\b", "", parte, flags=re.IGNORECASE)
            parte = re.sub(r"\be\b", "", parte, flags=re.IGNORECASE)
            parte = parte.strip()
            if parte:
                nomes_grupo.add(normalizar(parte))

        # encontrar o grupo cuja lista de membros tem maior sobreposição
        melhor_id, melhor_score = None, 0
        for gid, membros in membros_por_grupo.items():
            score = sum(1 for n in nomes_grupo
                        if any(n in m or m in n for m in membros))
            if score > melhor_score:
                melhor_id, melhor_score = gid, score

        if melhor_id is None or melhor_score == 0:
            continue

        questao = row[2]
        sim1, nao1 = row[3], row[4]
        questao_2 = row[7]
        estado = "Aprovado (1ª submissão)" if (sim1 and not nao1) else (
                 "Aprovado após reformulação" if questao_2 else
                 ("Para reformular" if nao1 else None))
        questao_final = questao_2 or questao

        db.execute("""UPDATE projeto_grupos SET questao_orientadora=?, estado_aprovacao=?
                      WHERE id=?""", (questao_final, estado, melhor_id))
        n_atualizados += 1

    db.commit()
    print(f"  → {n_atualizados} grupos com questão orientadora associada")
    return n_atualizados


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    args = sys.argv[1:]
    f_aval  = args[0] if len(args) > 0 else F_AVALIACAO
    f_grup  = args[1] if len(args) > 1 else F_GRUPOS
    f_quest = args[2] if len(args) > 2 else F_QUESTOES

    db = sqlite3.connect(DATABASE)
    db.row_factory = sqlite3.Row
    db.execute("PRAGMA foreign_keys = ON")

    if os.path.exists(f_aval):
        print(f"A importar avaliações de '{f_aval}'…")
        importar_avaliacoes(db, f_aval)
    else:
        print(f"! ficheiro não encontrado: {f_aval}")

    if os.path.exists(f_grup):
        print(f"A importar grupos/temas/orientadores de '{f_grup}'…")
        importar_grupos(db, f_grup)
    else:
        print(f"! ficheiro não encontrado: {f_grup}")

    if os.path.exists(f_quest):
        print(f"A importar questões orientadoras de '{f_quest}'…")
        importar_questoes(db, f_quest)
    else:
        print(f"! ficheiro não encontrado: {f_quest}")

    db.close()
    print("Concluído.")


if __name__ == "__main__":
    main()
